"""Full BOM export — the 12-sheet Oracle BOM deliverable.

Rebuilds the customer-facing BOM workbook (Table of Contents, Assumptions, Rate
Card, Pricing Overview, Compute, Storage, Networking, DR, Security KMS,
Consumption Ramp, Annexure, Applications Migrated to OCI) from the full-fidelity
build-spec (data/bom_template_spec.json), then populates its data sheets from the
app's priced inventory.

The spec carries every style, merge, image, data-validation, conditional format and
formula, so the rebuilt workbook is visually identical to the reference deliverable.
Only the DATA rows are replaced with the app's inventory.

Compute sheet contract (from the template):
    A VM/Server  B Tier  C Environment  D Master Application  E Master Description
    F Virtual/Physical  G OS Name  H OS Family  I vCPU/Cores  J OCPU (formula)
    K Memory GB  L Storage GB  M Block VPUs (formula)  N Monthly Hours
    O RAM $/mo  P OCPU $/mo  Q Block $/mo  R Total $/mo   (O..R formulas)
    Rate Card refs: C8=OCPU rate, C9=RAM rate, C10=block rate, C11=VPU rate,
    C12=default VPUs.  Compute!B9 = E6 optimization factor.
"""

import base64
import datetime
import io
import json
import re
import shutil
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import (AnchorMarker, OneCellAnchor,
                                                  TwoCellAnchor)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.formatting.rule import ColorScale, FormatObject, Rule
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation

SPEC_PATH = Path(__file__).resolve().parent / "data" / "bom_template_spec.json"

# openpyxl uses Pillow ONLY to read an image's width/height when embedding. That single
# dependency kept dropping the architecture diagram (and logos) on machines without Pillow.
# We don't need Pillow for that: PNG dimensions live in the file header. When Pillow is
# missing we parse the header ourselves and monkeypatch openpyxl so images embed anyway —
# the image bytes are written from the original file, never from PIL.
try:
    import PIL  # noqa: F401
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False


def _png_dimensions(src):
    """(width, height) from a PNG's IHDR header — no Pillow needed. src is a path or bytes."""
    import struct
    if hasattr(src, "read"):
        pos = src.tell(); head = src.read(24); src.seek(pos)
    elif isinstance(src, (bytes, bytearray)):
        head = bytes(src[:24])
    else:
        with open(src, "rb") as f:
            head = f.read(24)
    if head[:8] != b"\x89PNG\r\n\x1a\n" or head[12:16] != b"IHDR":
        return (0, 0)
    return struct.unpack(">II", head[16:24])


class _HeaderImage:
    """Stand-in for a PIL image. openpyxl's Image class only touches .size, .format, a
    readable .fp, and .close() — never any real decoding — so this is all it needs to embed
    a PNG. The raw bytes are carried through unchanged."""
    def __init__(self, data):
        import io
        self._bytes = bytes(data)
        self.size = _png_dimensions(self._bytes)
        self.format = "PNG"
        self.fp = io.BytesIO(self._bytes)

    def close(self):
        pass


def _enable_pillow_free_images():
    """Let openpyxl embed PNGs without Pillow by supplying dimensions from the header and
    handing back the raw bytes. openpyxl re-fetches the image via _import_image at write
    time, so this covers both metadata and the actual save. No-op when Pillow is present."""
    if HAS_PILLOW:
        return
    try:
        from openpyxl.drawing import image as _oxml_image
    except Exception:
        return

    def _import_image(img):
        if isinstance(img, _HeaderImage):
            return img
        if hasattr(img, "read"):
            pos = img.tell(); data = img.read(); img.seek(pos)
        else:
            with open(img, "rb") as f:
                data = f.read()
        return _HeaderImage(data)

    _oxml_image._import_image = _import_image
    # The constructor raises ImportError unless PILImage is truthy.
    if not getattr(_oxml_image, "PILImage", None):
        _oxml_image.PILImage = _HeaderImage


_enable_pillow_free_images()

COMPUTE_SHEET = "Compute"
COMPUTE_HEADER_ROW = 13
COMPUTE_FIRST_ROW = 14
COMPUTE_LAST_TEMPLATE_ROW = 689          # rows the reference deliverable ships with
COMPUTE_FORMULA_COLS = ["J", "M", "O", "P", "Q", "R"]
COMPUTE_ALL_COLS = list("ABCDEFGHIJKLMNOPQR")

APPS_SHEET = "Applications Migrated to OCI"
APPS_FIRST_ROW = 6
APPS_LAST_TEMPLATE_ROW = 207
APPS_FORMULA_COLS = list("BCDEFGH")

STORAGE_SHEET = "Storage"
STORAGE_FIRST_ROW = 10
STORAGE_LAST_TEMPLATE_ROW = 21


# ---------------------------------------------------------------------------
# Spec -> workbook (full-fidelity rebuild)
# ---------------------------------------------------------------------------
def _color(d):
    if not d:
        return None
    if "rgb" in d:
        return Color(rgb=d["rgb"])
    if "theme" in d:
        return Color(theme=d["theme"], tint=d.get("tint", 0.0))
    if "indexed" in d:
        return Color(indexed=d["indexed"])
    return None


def _font(d):
    if not d:
        return None
    return Font(name=d.get("name"), sz=d.get("size"), bold=d.get("bold", False),
                italic=d.get("italic", False), underline=d.get("underline"),
                strike=d.get("strike", False), color=_color(d.get("color")))


def _fill(d):
    if not d:
        return None
    return PatternFill(patternType=d["pattern"],
                       fgColor=_color(d.get("fgColor")) or Color(),
                       bgColor=_color(d.get("bgColor")) or Color())


def _side(d):
    return Side(style=d["style"], color=_color(d.get("color"))) if d else None


def _border(d):
    if not d:
        return None
    return Border(left=_side(d.get("left")), right=_side(d.get("right")),
                  top=_side(d.get("top")), bottom=_side(d.get("bottom")),
                  diagonal=_side(d.get("diagonal")),
                  diagonalUp=d.get("diagonalUp", False),
                  diagonalDown=d.get("diagonalDown", False))


def _align(d):
    if not d:
        return None
    return Alignment(horizontal=d.get("horizontal"), vertical=d.get("vertical"),
                     textRotation=d.get("textRotation", 0),
                     wrapText=d.get("wrapText", False),
                     shrinkToFit=d.get("shrinkToFit", False), indent=d.get("indent", 0))


def _unjval(v):
    if isinstance(v, dict) and "_type" in v:
        if v["_type"] == "datetime":
            return datetime.datetime.fromisoformat(v["iso"])
        if v["_type"] == "date":
            return datetime.date.fromisoformat(v["iso"])
        if v["_type"] == "time":
            return datetime.time.fromisoformat(v["iso"])
    return v


def load_spec():
    return json.loads(SPEC_PATH.read_text())


def build_workbook(spec):
    """Rebuild the template workbook from the build-spec (styles, images, formulas)."""
    styles = {}
    for sid, d in spec["styles"].items():
        styles[sid] = {
            "font": _font(d.get("font")), "fill": _fill(d.get("fill")),
            "border": _border(d.get("border")), "alignment": _align(d.get("alignment")),
            "number_format": d.get("number_format"),
        }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sd in spec["sheets"]:
        ws = wb.create_sheet(sd["name"])
        ws.sheet_state = sd.get("state", "visible")
        if sd.get("tab_color"):
            ws.sheet_properties.tabColor = _color(sd["tab_color"])
        v = sd.get("view", {})
        ws.sheet_view.showGridLines = v.get("showGridLines", True)
        ws.sheet_view.zoomScale = v.get("zoomScale", 100)
        if v.get("showRowColHeaders") is False:
            ws.sheet_view.showRowColHeaders = False
        if sd.get("freeze_panes"):
            ws.freeze_panes = sd["freeze_panes"]
        sf = sd.get("sheet_format", {})
        if sf.get("defaultColWidth"):
            ws.sheet_format.defaultColWidth = sf["defaultColWidth"]
        if sf.get("defaultRowHeight"):
            ws.sheet_format.defaultRowHeight = sf["defaultRowHeight"]

        for col, d in sd.get("column_dimensions", {}).items():
            cd = ws.column_dimensions[col]
            if "width" in d:
                cd.width = d["width"]
            if d.get("hidden"):
                cd.hidden = True
            if d.get("outlineLevel"):
                cd.outlineLevel = d["outlineLevel"]
            if "range" in d:
                cd.min, cd.max = d["range"]
        for row, d in sd.get("row_dimensions", {}).items():
            rd = ws.row_dimensions[int(row)]
            if "height" in d:
                rd.height = d["height"]
            if d.get("hidden"):
                rd.hidden = True
            if d.get("outlineLevel"):
                rd.outlineLevel = d["outlineLevel"]

        # Merges first: merge_cells() wipes styles set inside the range.
        for rng in sd.get("merged_cells", []):
            ws.merge_cells(rng)

        for coord, cd in sd["cells"].items():
            c = ws[coord]
            if "f" in cd:
                c.value = cd["f"]
            elif "v" in cd:
                c.value = _unjval(cd["v"])
            if "s" in cd:
                st = styles[cd["s"]]
                if st["font"]:
                    c.font = st["font"]
                if st["fill"]:
                    c.fill = st["fill"]
                if st["border"]:
                    c.border = st["border"]
                if st["alignment"]:
                    c.alignment = st["alignment"]
                if st["number_format"]:
                    c.number_format = st["number_format"]
            if "hyperlink" in cd:
                from openpyxl.worksheet.hyperlink import Hyperlink
                hl = cd["hyperlink"]
                c.hyperlink = Hyperlink(ref=coord, target=hl.get("target"),
                                        location=hl.get("location"), tooltip=hl.get("tooltip"))
            if "comment" in cd:
                c.comment = Comment(cd["comment"]["text"], cd["comment"].get("author") or "")

        if sd.get("auto_filter"):
            ws.auto_filter.ref = sd["auto_filter"]

        for r in sd.get("conditional_formatting", []):
            kw = {"type": r["type"]}
            if "operator" in r:
                kw["operator"] = r["operator"]
            if "formula" in r:
                kw["formula"] = r["formula"]
            if "priority" in r:
                kw["priority"] = r["priority"]
            if r.get("stopIfTrue"):
                kw["stopIfTrue"] = True
            if "dxf" in r:
                dx = r["dxf"]
                kw["dxf"] = DifferentialStyle(font=_font(dx.get("font")),
                                              fill=_fill(dx.get("fill")),
                                              border=_border(dx.get("border")))
            if "colorScale" in r:
                cs = r["colorScale"]
                kw["colorScale"] = ColorScale(
                    cfvo=[FormatObject(type=o["type"], val=o["val"]) for o in cs["cfvo"]],
                    color=[_color(c) for c in cs["colors"]])
            ws.conditional_formatting.add(r["range"], Rule(**kw))

        for d in sd.get("data_validations", []):
            dv = DataValidation(type=d.get("type"), operator=d.get("operator"),
                                formula1=d.get("formula1"), formula2=d.get("formula2"),
                                allowBlank=d.get("allowBlank", False),
                                showInputMessage=d.get("showInputMessage", False),
                                showErrorMessage=d.get("showErrorMessage", False),
                                errorTitle=d.get("errorTitle"), error=d.get("error"),
                                promptTitle=d.get("promptTitle"), prompt=d.get("prompt"))
            for rng in d["ranges"].split():
                dv.add(rng)
            ws.add_data_validation(dv)

        for im in sd.get("images", []):
            data = base64.b64decode(spec["images"][im["image_ref"]]["base64"])
            img = XLImage(io.BytesIO(data))
            a = im["anchor"]
            fr = a["from"]
            m1 = AnchorMarker(col=fr["col"], colOff=fr["colOff"], row=fr["row"], rowOff=fr["rowOff"])
            if a["type"] == "TwoCellAnchor" and "to" in a:
                t = a["to"]
                m2 = AnchorMarker(col=t["col"], colOff=t["colOff"], row=t["row"], rowOff=t["rowOff"])
                img.anchor = TwoCellAnchor(_from=m1, to=m2)
            else:
                ext = a.get("ext_emu")
                img.anchor = OneCellAnchor(
                    _from=m1, ext=XDRPositiveSize2D(cx=ext["cx"], cy=ext["cy"]) if ext else None)
            ws.add_image(img)

    order = spec["workbook"]["sheet_order"]
    # Open on the front of the deliverable, not wherever the source workbook was last saved
    # (the template had "Compute" active, so the export opened mid-way through the BOM).
    landing = "Table of Contents" if "Table of Contents" in order else order[0]
    wb.active = order.index(landing)
    for i, ws in enumerate(wb.worksheets):
        ws.sheet_view.tabSelected = (i == wb.active)
    return wb


def _postprocess(path, spec):
    """Restore the theme, the default font, and zero-height rows (openpyxl drops them)."""
    wbk = spec["workbook"]

    def font_xml(d):
        parts = [f'<name val="{d["name"]}"/>']
        if "size" in d:
            parts.append(f'<sz val="{d["size"]:g}"/>')
        if d.get("bold"):
            parts.append("<b/>")
        if d.get("italic"):
            parts.append("<i/>")
        c = d.get("color")
        if c and "rgb" in c:
            parts.append(f'<color rgb="{c["rgb"]}"/>')
        elif c and "theme" in c:
            parts.append(f'<color theme="{c["theme"]}"/>')
        return "<font>" + "".join(parts) + "</font>"

    zero_rows = {}
    for i, sd in enumerate(spec["sheets"], start=1):
        zr = [r for r, d in sd.get("row_dimensions", {}).items() if d.get("height") == 0]
        if zr:
            zero_rows[f"xl/worksheets/sheet{i}.xml"] = zr

    tmp = str(path) + ".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/theme/theme1.xml" and wbk.get("theme_xml_base64"):
                data = base64.b64decode(wbk["theme_xml_base64"])
            elif item.filename == "xl/styles.xml" and wbk.get("default_font"):
                txt = data.decode("utf-8")
                txt = re.sub(r"(<fonts[^>]*>)<font>.*?</font>",
                             lambda m: m.group(1) + font_xml(wbk["default_font"]), txt, count=1)
                data = txt.encode("utf-8")
            elif item.filename in zero_rows:
                txt = data.decode("utf-8")
                for r in zero_rows[item.filename]:
                    m = re.search(f'<row r="{r}"[^>]*>', txt)
                    if not m:
                        continue
                    tag = m.group(0)
                    if ' ht="' not in tag:
                        new = tag.replace(f'<row r="{r}"', f'<row r="{r}" ht="0"', 1)
                        if "customHeight=" not in new:
                            new = new.replace(f'<row r="{r}"', f'<row r="{r}" customHeight="1"', 1)
                        txt = txt.replace(tag, new, 1)
                data = txt.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, path)


# ---------------------------------------------------------------------------
# Field resolution from the app's uploaded inventory
# ---------------------------------------------------------------------------
def _norm(v):
    return re.sub(r"[^a-z0-9]+", " ", str(v or "").lower()).strip()


def _clean(v):
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    return "" if s.lower() in {"nan", "none", "nat"} else s


def _field_source_text(f):
    """The field's ORIGINAL column header. The app renames CPU/memory columns to
    'Application Details: OCPUs' / '... Memory per server (GB)', so matching on the
    renamed label would wrongly resolve e.g. 'Application' to the OCPU column."""
    original = f.get("cpuSourceLabel") or f.get("memorySourceLabel") or f.get("sourceHeader")
    return _norm(original or f.get("label"))


def _find_field(fields, *needle_sets):
    """First field whose ORIGINAL header matches any needle set (all terms present)."""
    for needles in needle_sets:
        for f in fields or []:
            if not isinstance(f, dict):
                continue
            text = _field_source_text(f)
            if all(n in text for n in needles):
                return f.get("key")
    return None


def _distinct_sites(fields, rows):
    """How many distinct sites the inventory actually names — None when it has no site
    column at all (most VM exports don't). Drives whether the diagram draws remote sites."""
    key = _find_field(fields, ["site"], ["location"], ["datacenter"], ["data", "center"],
                      ["campus"], ["facility"])
    if not key:
        return None
    vals = {_clean(r.get(key)) for r in (rows or []) if _clean(r.get(key))}
    return len(vals) or None


def _resolve_inventory_keys(fields):
    return {
        "server": _find_field(fields, ["server", "name"], ["vm", "name"], ["host", "name"], ["machine"]),
        "os_name": _find_field(fields, ["guest", "os"], ["os", "name"], ["operating", "system"]),
        "os_family": _find_field(fields, ["os", "type"], ["os", "family"], ["platform"]),
        "virt": _find_field(fields, ["physical"], ["virtual"], ["server", "type"]),
        "env": _find_field(fields, ["environment"], ["env"]),
        "app": _find_field(fields, ["application", "name"], ["application"], ["app"]),
        "tier": _find_field(fields, ["tier"]),
        "desc": _find_field(fields, ["description"]),
    }


# ---------------------------------------------------------------------------
# Populate
# ---------------------------------------------------------------------------
def _clear_range(ws, first_row, last_row, cols):
    for r in range(first_row, last_row + 1):
        for c in cols:
            ws[f"{c}{r}"] = None


def _populate_compute(ws, servers, hours, rate_refs=None, shape_label=""):
    """Write the app's servers into the Compute sheet, translating the template's
    row-14 formulas down to each row and clearing any unused template rows. The live
    OCPU/RAM/block/VPU formulas point at the Rate Card cells that the rebuilt (used-only)
    rate card actually placed those rates on, so the math is transparent."""
    rate_refs = rate_refs or {}
    protos = {c: ws[f"{c}{COMPUTE_FIRST_ROW}"].value for c in COMPUTE_FORMULA_COLS}
    proto_styles = {c: ws[f"{c}{COMPUTE_FIRST_ROW}"]._style for c in COMPUTE_ALL_COLS}
    # The app prices each line item to the cent and then sums. Mirror that rounding here
    # (same rates/refs, just ROUND-ed) so the workbook ties out to the app to the penny
    # instead of drifting by fractions of a cent across hundreds of rows.
    R = COMPUTE_FIRST_ROW
    ram_ref = rate_refs.get("ram")
    ocpu_ref = rate_refs.get("ocpu")
    block_ref = rate_refs.get("block")
    vpu_ref = rate_refs.get("vpu")
    vpus_ref = rate_refs.get("vpus")
    protos["O"] = (f"=ROUND(K{R}*N{R}*'Rate Card'!$C${ram_ref},2)" if ram_ref else 0)
    # Compute optimization adjusts the OCPU/RAM QUANTITIES (columns J and K) — it must NOT
    # also discount the price here, or it would double-count. Price flows straight from the
    # optimized sizing.
    protos["P"] = (f"=ROUND(J{R}*N{R}*'Rate Card'!$C${ocpu_ref},2)" if ocpu_ref else 0)
    if block_ref and vpu_ref:
        protos["Q"] = (f"=IF($L{R}=\"\",0,ROUND($L{R}*'Rate Card'!$C${block_ref},2)"
                       f"+ROUND($L{R}*'Rate Card'!$C${vpu_ref}*IF($M{R}=\"\",0,$M{R}),2))")
    else:
        protos["Q"] = 0
    # Block-VPU seed (column M) references the Rate Card default-VPUs cell.
    protos["M"] = (f"=IF($A{R}<>\"\",'Rate Card'!$C${vpus_ref},\"\")" if vpus_ref else '=""')

    # Column S states the OCI shape each server is mapped to.
    ws.cell(13, 19).value = "OCI Shape (Mapped)"
    ws.cell(13, 19)._style = ws.cell(13, 1)._style
    # The O/P/R headers hard-coded "E6"; restate them for whatever shape is used.
    if shape_label:
        ws["O13"] = f"{shape_label} RAM Monthly"
        ws["P13"] = f"{shape_label} OCPU Monthly"
        ws["R13"] = f"{shape_label} Total Monthly"

    last_written = COMPUTE_FIRST_ROW - 1
    for i, s in enumerate(servers):
        r = COMPUTE_FIRST_ROW + i
        # Extend styling past the template's shipped rows if the inventory is bigger.
        if r > COMPUTE_LAST_TEMPLATE_ROW:
            for c in COMPUTE_ALL_COLS:
                ws[f"{c}{r}"]._style = proto_styles[c]
        ws[f"A{r}"] = s.get("server") or s.get("app") or f"Server {i + 1}"
        ws[f"B{r}"] = s.get("tier") or None
        ws[f"C{r}"] = s.get("env") or None
        ws[f"D{r}"] = s.get("app") or None
        ws[f"E{r}"] = s.get("desc") or None
        ws[f"F{r}"] = s.get("virt") or None
        ws[f"G{r}"] = s.get("os_name") or None
        ws[f"H{r}"] = s.get("os_family") or None
        ws[f"I{r}"] = s.get("vcpu") or None
        ws[f"K{r}"] = s.get("memory_gb") or None
        ws[f"L{r}"] = s.get("storage_gb") or None
        # Per-row hours from the data source (falls back to the global hours).
        ws[f"N{r}"] = s.get("hours") or hours
        ws[f"S{r}"] = s.get("shape") or shape_label or None
        for c in COMPUTE_FORMULA_COLS:
            proto = protos.get(c)
            if proto:
                ws[f"{c}{r}"] = Translator(
                    proto, origin=f"{c}{COMPUTE_FIRST_ROW}").translate_formula(f"{c}{r}")
        # The source workbook shipped its data rows COLLAPSED (hidden=True on all 1,061 of
        # them), and the spec reproduces that faithfully — so every server we wrote was
        # invisible in the export and the sheet looked empty. Show the rows we populate.
        rd = ws.row_dimensions[r]
        rd.hidden = False
        if not rd.height:
            rd.height = 28.0
        last_written = r

    # Blank out any leftover reference rows so SUBTOTAL/COUNTA don't pick them up, and
    # collapse them so the sheet ends at the last real server instead of trailing a
    # thousand empty rows.
    if last_written < COMPUTE_LAST_TEMPLATE_ROW:
        _clear_range(ws, last_written + 1, COMPUTE_LAST_TEMPLATE_ROW, COMPUTE_ALL_COLS)
        for r in range(last_written + 1, COMPUTE_LAST_TEMPLATE_ROW + 1):
            ws.row_dimensions[r].hidden = True
    return last_written


def _populate_apps(ws, apps):
    """Applications Migrated to OCI: app names in col A; B..H are template formulas."""
    protos = {c: ws[f"{c}{APPS_FIRST_ROW}"].value for c in APPS_FORMULA_COLS}
    last = APPS_FIRST_ROW - 1
    for i, name in enumerate(apps):
        r = APPS_FIRST_ROW + i
        if r > APPS_LAST_TEMPLATE_ROW:
            break
        ws[f"A{r}"] = name
        for c in APPS_FORMULA_COLS:
            proto = protos.get(c)
            if proto:
                ws[f"{c}{r}"] = Translator(
                    proto, origin=f"{c}{APPS_FIRST_ROW}").translate_formula(f"{c}{r}")
        last = r
    if last < APPS_LAST_TEMPLATE_ROW:
        _clear_range(ws, last + 1, APPS_LAST_TEMPLATE_ROW, list("ABCDEFGH"))


def _populate_storage(ws, storage_rows, rate_refs=None, file_rate=None):
    """Storage sheet: object/file-storage candidates (cleared when the app has none).

    Each row can supply either (gb, rate) -> I = gb*rate, or a direct `monthly` value when
    capacity isn't known (cloud-bill storage is usage-priced, so GB isn't always available).
    File-storage rows point their rate cell (H) at the Rate Card file-storage rate for
    transparency."""
    rate_refs = rate_refs or {}
    file_ref = rate_refs.get("file")
    proto = ws[f"A{STORAGE_FIRST_ROW}"]._style
    _clear_range(ws, STORAGE_FIRST_ROW, STORAGE_LAST_TEMPLATE_ROW, list("ABCDEFGHIJ"))
    for i, s in enumerate(storage_rows):
        r = STORAGE_FIRST_ROW + i
        if r > STORAGE_LAST_TEMPLATE_ROW:
            for c in "ABCDEFGHIJ":
                ws[f"{c}{r}"]._style = proto
        ws[f"A{r}"] = s.get("server")
        ws[f"B{r}"] = s.get("tier")
        ws[f"C{r}"] = s.get("env")
        ws[f"D{r}"] = s.get("app")
        ws[f"E{r}"] = s.get("signal")
        ws[f"F{r}"] = s.get("target")
        ws[f"G{r}"] = s.get("gb")
        rate = s.get("rate")
        if s.get("gb") and rate:
            if file_ref and file_rate and abs(float(rate) - float(file_rate)) < 1e-9:
                ws[f"H{r}"] = f"='Rate Card'!$C${file_ref}"
            else:
                ws[f"H{r}"] = rate
            ws[f"I{r}"] = f"=G{r}*H{r}"
        else:
            ws[f"H{r}"] = rate
            ws[f"I{r}"] = round(float(s.get("monthly") or 0), 2)


def _cloud_storage_rows(pricing):
    """Aggregate the cloud-bill Storage-category services by OCI product for the Storage
    sheet, so it itemizes what rolled into the Pricing Overview Storage line."""
    import bom_export
    agg = {}
    for r in (pricing or {}).get("rows", []):
        if (r.get("costAction") or "") == "remove":
            continue
        if bom_export._cloud_product_group(r.get("ociServiceCategory"),
                                           r.get("sourceService")) != "Storage":
            continue
        prod = _clean(r.get("ociProduct")) or "OCI Storage"
        specs = r.get("specs") or {}
        gb = (float(specs.get("blockStorageGb") or 0) + float(specs.get("fileStorageGb") or 0)
              + float(specs.get("cloudStorageGb") or 0))
        a = agg.setdefault(prod, {"gb": 0.0, "monthly": 0.0, "svc": _clean(r.get("sourceService"))})
        a["gb"] += gb
        a["monthly"] += float(r.get("monthly") or 0)
    rows = []
    for prod, v in sorted(agg.items(), key=lambda kv: -kv[1]["monthly"]):
        if v["monthly"] <= 0:
            continue
        rows.append({
            "server": prod, "signal": "Mapped from cloud bill",
            "target": prod, "app": v["svc"],
            "gb": round(v["gb"], 2) or None,
            "rate": round(v["monthly"] / v["gb"], 6) if v["gb"] else None,
            "monthly": round(v["monthly"], 2),
        })
    return rows


RATE_CARD_HDR_ROW = 7      # header row for the rate table
RATE_CARD_FIRST_ROW = 8    # first data row
RATE_CARD_CLEAR_LAST = 60  # scrub the template's fixed sections well past the last row


def _collect_rate_card_entries(shape, block_rate, vpu_rate, default_vpus, hours,
                               file_rate, windows_rate, windows_sku, windows_priced,
                               servers, storage_rows, pricing, extra_services,
                               is_cloud_bill):
    """Build the list of rate-card lines that were ACTUALLY used in this build. Core
    compute/storage/licensing inputs (carrying a `key` so the sheets can reference them),
    plus every distinct mapped-service SKU that appears in the priced line items and any
    user-added OCI services. Returns a list of {sku, name, val, unit, note, key}."""
    entries = []

    def add(sku, name, val, unit, note, key=None):
        entries.append({"sku": (sku or "N/A"), "name": name, "val": val,
                        "unit": unit, "note": note, "key": key})

    has_compute = bool(servers)
    has_block = any(s.get("storage_gb") for s in servers)
    has_file = any("file" in (s.get("target") or "").lower() for s in (storage_rows or []))

    label = (shape or {}).get("shortLabel") or (shape or {}).get("label") or "OCI"
    if has_compute and shape:
        if shape.get("computeRate") is not None:
            add(shape.get("computeSku"), f"{label} OCPU per hour", float(shape["computeRate"]),
                "per OCPU-hour", f"OCI {label} Flex Compute OCPU pricing (Compute col P).", "ocpu")
        if shape.get("memoryRate") is not None:
            add(shape.get("memorySku"), f"{label} RAM GB per hour", float(shape["memoryRate"]),
                "per GB-hour", f"OCI {label} Flex Compute memory pricing (Compute col O).", "ram")
    if has_block:
        add("B91961", "VM Block Volume Storage", float(block_rate), "per GB-month",
            "VM-attached block storage (Compute col Q).", "block")
        add("B91962", "VM Block Volume Performance Units", float(vpu_rate), "per GB-month",
            "Block Volume performance (VPU) component (Compute col Q).", "vpu")
        add("N/A", "Default VM Block Volume VPUs", float(default_vpus), "VPUs / GB-month",
            "Default performance units seeded into the Compute VPU column (col M).", "vpus")
    if has_compute:
        add("N/A", "Monthly hours", float(hours), "hours / month",
            "Default monthly-hours assumption (Compute col N).", "hours")
        if not is_cloud_bill:
            add("N/A", "vCPU / core to OCPU conversion",
                "Virtual: 2 vCPU = 1 OCPU; Physical: 1 core = 1 OCPU", "conversion",
                "Used when deriving OCPUs from the source inventory.", "conv")
    if has_file:
        add("B89057", "File Storage Service", float(file_rate), "per GB-month",
            "File / NAS storage rows (Storage sheet).", "file")
    if windows_priced and windows_rate is not None:
        add(windows_sku, "Windows OS licensing", float(windows_rate), "per OCPU-hour",
            "Windows rows; shown as 3rd Party Licensing on the Pricing Overview.", "windows")

    # Every distinct mapped-service SKU that appears in the priced line items (covers
    # cloud-bill services: OIC, ADW, Load Balancer, WAF, DNS, Object Storage, etc.). A SKU
    # can appear on many lines (e.g. OIC's single priced anchor plus zero-cost consolidated
    # rows) — keep the line with the highest unit rate so the rate card shows the real rate.
    core_skus = {e["sku"] for e in entries if e["sku"] and e["sku"] != "N/A"}
    info = {}

    def note_used(sku, name, unit, val):
        if not sku or sku in core_skus:
            return
        val = val if isinstance(val, (int, float)) else None
        cur = info.get(sku)
        if cur is None:
            info[sku] = {"name": name or sku, "unit": unit or "", "val": val}
        elif (val or 0) > (cur["val"] or 0):
            cur.update(name=name or cur["name"], unit=unit or cur["unit"], val=val)

    for r in (pricing or {}).get("rows", []):
        if (r.get("costAction") or "") == "remove":
            continue
        for li in (r.get("lineItems") or []):
            note_used(_clean(li.get("sku")), _clean(li.get("description")),
                      _clean(li.get("unit")), li.get("rate"))

    if extra_services:
        try:
            import oci_catalog
            priced, _ = oci_catalog.price_extras(extra_services, hours)
            for e in priced:
                note_used(_clean(e.get("sku")), _clean(e.get("name")), _clean(e.get("unit")), None)
        except Exception:
            pass

    for sku, v in info.items():
        add(sku, v["name"], v["val"], v["unit"], "Mapped-service rate used in this build.")

    return entries


def _write_rate_card(ws, entries):
    """Rewrite the Rate Card sheet with only the used entries, sorted alphabetically by
    SKU (N/A inputs after the real SKUs, by name). Returns {key: row} so the other sheets
    can point their live formulas at the exact cells for transparency."""
    # Preserve the template's header + data cell styling, then scrub its fixed sections.
    cols = "ABCDE"
    hdr_style = {c: ws.cell(RATE_CARD_HDR_ROW, i)._style for i, c in enumerate(cols, 1)}
    data_style = {c: ws.cell(RATE_CARD_FIRST_ROW, i)._style for i, c in enumerate(cols, 1)}
    # The template's section headers span merged cells; unmerge anything in the scrub
    # region so we can overwrite it, then blank it out.
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 6 and mr.max_row <= RATE_CARD_CLEAR_LAST:
            ws.unmerge_cells(str(mr))
    for r in range(6, RATE_CARD_CLEAR_LAST + 1):
        for i in range(1, 7):
            ws.cell(r, i).value = None

    ws.cell(4, 1).value = "RATE CARD — ONLY THE SKUS / RATES USED IN THIS BUILD"
    headers = ["SKU", "Input", "Value", "Unit", "Workbook Use / Note"]
    for i, (c, h) in enumerate(zip(cols, headers), 1):
        cell = ws.cell(RATE_CARD_HDR_ROW, i)
        cell.value = h
        cell._style = hdr_style[c]

    def sort_key(e):
        sku = e["sku"]
        if sku and sku != "N/A":
            return (0, sku.upper(), e["name"].upper())
        return (1, "", e["name"].upper())

    refs = {}
    for idx, e in enumerate(sorted(entries, key=sort_key)):
        r = RATE_CARD_FIRST_ROW + idx
        vals = [e["sku"], e["name"], e["val"], e["unit"], e["note"]]
        for i, (c, v) in enumerate(zip(cols, vals), 1):
            cell = ws.cell(r, i)
            cell.value = v
            cell._style = data_style[c]
        cval = e["val"]
        if isinstance(cval, (int, float)):
            ws.cell(r, 3).number_format = "#,##0" if float(cval).is_integer() else "0.0000"
        if e.get("key"):
            refs[e["key"]] = r
    return refs


def embed_architecture(ws_po, png_path, anchor_spec=None):
    """Drop this BOM's generated diagram into the Pricing Overview's architecture slot.

    The template no longer carries an architecture picture — only its ANCHOR (the cell
    footprint the picture occupied), kept in the spec as `architecture_anchor`. The image
    itself was the source workbook's own architecture drawing: 637 KB of another customer's
    diagram, decoded and inserted on every export just to be overwritten. Storing the
    footprint and nothing else means an export with no diagram simply has no picture, and
    one with a diagram gets ours in exactly the right place.
    """
    if not png_path or not Path(png_path).exists() or not anchor_spec:
        return False

    fr = anchor_spec["from"]
    m1 = AnchorMarker(col=fr["col"], colOff=fr["colOff"], row=fr["row"], rowOff=fr["rowOff"])
    new = XLImage(str(png_path))
    if anchor_spec.get("type") == "TwoCellAnchor" and "to" in anchor_spec:
        t = anchor_spec["to"]
        m2 = AnchorMarker(col=t["col"], colOff=t["colOff"], row=t["row"], rowOff=t["rowOff"])
        new.anchor = TwoCellAnchor(_from=m1, to=m2)
    else:
        ext = anchor_spec.get("ext_emu")
        new.anchor = OneCellAnchor(
            _from=m1, ext=XDRPositiveSize2D(cx=ext["cx"], cy=ext["cy"]) if ext else None)
    ws_po.add_image(new)
    return True


CUSTOMER_TOKEN = "{{CUSTOMER}}"


def _apply_customer_name(wb, bom_name):
    """Swap the template's {{CUSTOMER}} token for the BOM name the user typed.

    The template is deliberately customer-neutral — it must never ship one client's name
    (or servers, or numbers) inside another client's deliverable. Every customer-specific
    string in the spec is the token; this is the only place a real name enters the
    workbook. With no BOM name, it degrades to a generic "the customer".
    """
    name = (bom_name or "").strip()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and CUSTOMER_TOKEN in v:
                    if not name:
                        # Mid-sentence the generic reads "the customer"; as a title it
                        # would read "The customer OCI Migration", so trim the token out.
                        v = (v.replace(CUSTOMER_TOKEN + " ", "the customer ")
                             if not v.startswith(CUSTOMER_TOKEN)
                             else v[len(CUSTOMER_TOKEN):].lstrip())
                        v = v.replace(CUSTOMER_TOKEN, "the customer")
                    else:
                        v = v.replace(CUSTOMER_TOKEN, name)
                    cell.value = v


def _set_toc(ws, bom_name):
    if bom_name:
        ws["B5"] = bom_name
        ws["A1"] = f"{bom_name} — OCI Migration BOM"
    ws["B8"] = datetime.date.today().isoformat()


def _clear_cell_note(ws, coord):
    """Strip the hidden tooltip on a cell: a hover comment/note AND any data-validation
    input prompt (Excel shows a DV prompt as a note when the cell is selected). B9 shipped
    a stale prompt titled 'E6 Optimization' that no longer describes what the cell does."""
    try:
        ws[coord].comment = None
    except Exception:
        pass
    try:
        dvs = list(ws.data_validations.dataValidation)
    except Exception:
        return
    keep = []
    for dv in dvs:
        sq = str(dv.sqref).strip()
        if coord in sq:
            if sq == coord:
                continue                    # covers only this cell -> drop entirely
            dv.showInputMessage = False     # covers more -> just silence the prompt
            dv.prompt = None
            dv.promptTitle = None
        keep.append(dv)
    ws.data_validations.dataValidation = keep


# Cloud-bill service groups -> Pricing Overview line (same grouping the Product Breakdown
# uses). Compute stays in the Compute sheet (lines 13-15). Everything else rolls into the
# matching Overview line so the template total includes the whole bill, not just compute.
_CLOUD_GROUP_TO_OVERVIEW_ROW = {
    "Storage": 16, "Networking": 18, "Security": 19,
    "Database": 16, "Obs. & Management": 16, "Other Services": 16,
    "AI & Machine Learning": 16, "DevOps": 16, "Marketplace": 16, "Support": 16,
}


def _cloud_effective_hours(pr):
    """Effective monthly hours for a cloud-bill compute row = billed OCPU-hours / OCPUs.
    The bill meters actual usage, so a line can be far more (or less) than 730 hours; using
    the effective hours makes the Compute sheet's OCPU x hours x rate reproduce the app's
    actual cost. Falls back to the row's hoursPerMonth when the OCPU-hr quantity is absent."""
    specs = pr.get("specs") or {}
    ocpus = float(specs.get("ocpus") or 0)
    if ocpus > 0:
        for li in (pr.get("lineItems") or []):
            d = (li.get("description") or "").lower()
            if "ocpu" in d and ("hr" in d or "hour" in d):
                q = float(li.get("quantity") or 0)
                if q > 0:
                    return round(q / ocpus, 4)
    return float(pr.get("hoursPerMonth") or 0) or None


def _add_cloud_bill_services(wb, pricing):
    """Set the Pricing Overview lines from the app's EXACT per-row monthly so the whole
    cloud bill ties out — every service, not just compute.

    Each bill line's monthly is allocated to one Overview line by category. Compute rows
    carry non-OCPU/RAM costs too (attached storage, data transfer...), so compute is split
    into its OCPU / RAM / other-compute line-item sums. The Overview lines are written as
    VALUES (replacing the =SUM(Compute!...) formulas) because the Compute sheet re-derives
    from spec x hours and can't reproduce a metered bill's true totals. Sum of lines 13-20
    then equals the app's OCI monthly exactly; 3rd-party licensing stays on line 21."""
    import bom_export
    ws = wb["Pricing Overview"]
    comp_ocpu = comp_ram = comp_other = 0.0
    svc = {}                       # overview row -> total
    other_non_storage = False
    for r in (pricing or {}).get("rows", []):
        if (r.get("costAction") or "") == "remove":
            continue
        grp = bom_export._cloud_product_group(r.get("ociServiceCategory"), r.get("sourceService"))
        if grp == "Compute":
            for li in (r.get("lineItems") or []):
                d = (li.get("description") or "").lower()
                m = float(li.get("monthly") or 0)
                if "ocpu" in d and ("hr" in d or "hour" in d):
                    comp_ocpu += m
                elif "memory" in d and ("hr" in d or "hour" in d):
                    comp_ram += m
                else:
                    comp_other += m
        else:
            row = _CLOUD_GROUP_TO_OVERVIEW_ROW.get(grp, 16)
            svc[row] = svc.get(row, 0.0) + float(r.get("monthly") or 0)
            if row == 16 and grp != "Storage":
                other_non_storage = True

    # Compute lines: OCPU -> 13, RAM -> 14, everything else on compute rows -> 15.
    ws["B13"] = round(comp_ocpu, 2)
    ws["B14"] = round(comp_ram, 2)
    ws["B15"] = round(comp_other, 2)
    if comp_other:
        ws["A15"] = "VM Block Storage + attached:"
    # Service lines by group.
    for row, amt in svc.items():
        base = ws[f"B{row}"].value
        base = 0.0 if isinstance(base, str) else float(base or 0)
        ws[f"B{row}"] = round(base + amt, 2)
    if 16 in svc and other_non_storage:
        ws["A16"] = "Storage / Other OCI Services:"


# The 11 product groups (matches the Product Breakdown) and the template sheets that
# already cover four of them; the rest get a generated detail sheet.
_PRODUCT_GROUPS = [
    "Compute", "Database", "Storage", "Networking", "Support",
    "Obs. & Management", "Other Services", "Security",
    "AI & Machine Learning", "DevOps", "Marketplace",
]
_GROUP_EXISTING_SHEET = {
    "Compute": "Compute", "Storage": "Storage",
    "Networking": "Networking", "Security": "Security KMS",
}
_GROUP_SHEET_NAME = {
    "Obs. & Management": "Obs. and Management",
    "AI & Machine Learning": "AI and Machine Learning",
}


def _aggregate_product_groups(pricing):
    """Group the priced rows into the 11 product groups, returning
    {group: {"aws", "oci", "items": {(awsService, ociProduct): {"aws","oci"}}}}."""
    import bom_export
    groups = {}
    for r in (pricing or {}).get("rows", []):
        if (r.get("costAction") or "") == "remove":
            continue
        grp = bom_export._cloud_product_group(
            r.get("ociServiceCategory"), r.get("sourceService") or "Other",
            r.get("ociProduct"))
        aws = float(r.get("sourceMonthlyCost") or 0)
        oci = float(r.get("monthly") or 0)
        g = groups.setdefault(grp, {"aws": 0.0, "oci": 0.0, "items": {}})
        g["aws"] += aws
        g["oci"] += oci
        key = (_clean(r.get("sourceService")) or "Other",
               _clean(r.get("ociProduct")) or "Needs review")
        it = g["items"].setdefault(key, {"aws": 0.0, "oci": 0.0})
        it["aws"] += aws
        it["oci"] += oci
    return groups


def _add_product_group_topics(wb, pricing):
    """Cloud-bill: put a 'Cost by Product Group' summary of ALL 11 topics on the Pricing
    Overview, and add a detail sheet for every group that has cost and doesn't already have
    a dedicated sheet (Compute/Storage/Networking/Security KMS already exist)."""
    import bom_export
    groups = _aggregate_product_groups(pricing)
    ctr = Alignment(horizontal="center", vertical="center")
    hdr_font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
    lbl_font = Font(name="Calibri", size=11, bold=True)
    money_fmt = "#,##0.00"

    # ---- per-group detail sheets for groups without a dedicated sheet ----
    for grp in _PRODUCT_GROUPS:
        g = groups.get(grp)
        if not g or (g["oci"] <= 0 and g["aws"] <= 0):
            continue
        if grp in _GROUP_EXISTING_SHEET:
            continue
        name = (_GROUP_SHEET_NAME.get(grp, grp))[:31]
        if name in wb.sheetnames:
            continue
        ws2 = wb.create_sheet(name)
        ws2.column_dimensions["A"].width = 46
        ws2.column_dimensions["B"].width = 46
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 18
        ws2["A1"] = f"{grp} — OCI mapped services"
        ws2["A1"].font = Font(name="Calibri", size=14, bold=True)
        for c, txt in ((1, "AWS Service"), (2, "OCI Product"),
                       (3, "AWS Monthly"), (4, "OCI Monthly")):
            cell = ws2.cell(3, c)
            cell.value = txt
            cell.font = hdr_font
            cell.fill = PatternFill("solid", fgColor="FF4472C4")
            cell.alignment = ctr
        rr = 4
        for (aws_svc, oci_prod), v in sorted(g["items"].items(), key=lambda kv: -kv[1]["oci"]):
            ws2.cell(rr, 1).value = aws_svc
            ws2.cell(rr, 2).value = oci_prod
            ws2.cell(rr, 3).value = round(v["aws"], 2)
            ws2.cell(rr, 4).value = round(v["oci"], 2)
            ws2.cell(rr, 3).number_format = money_fmt
            ws2.cell(rr, 4).number_format = money_fmt
            rr += 1
        ws2.cell(rr, 1).value = "Total"
        ws2.cell(rr, 1).font = lbl_font
        ws2.cell(rr, 3).value = f"=SUM(C4:C{rr-1})"
        ws2.cell(rr, 4).value = f"=SUM(D4:D{rr-1})"
        ws2.cell(rr, 3).number_format = money_fmt
        ws2.cell(rr, 4).number_format = money_fmt
        ws2.cell(rr, 3).font = lbl_font
        ws2.cell(rr, 4).font = lbl_font


def _set_optimization(ws_compute, rightsized=False, ocpu_pct=0.0, ram_pct=0.0, is_ax=False):
    """Relabel and fill the Compute-optimization block.

    The optimization does NOT discount the price directly — it shrinks each VM's OCPU and
    RAM quantities (columns J/K) via ceil(value*(1-pct)), floored at 2, and the price then
    flows from the smaller sizing. B9 is a record of what was applied ("% approximation"),
    not a live multiplier. For Ax shapes the base %% deepens with the source instance's
    generation gap (x2 per generation behind the first), so each row can differ.
    """
    ws_compute["A8"] = "Compute optimization"
    ws_compute["A9"] = "% approximation"
    if rightsized and (ocpu_pct or ram_pct):
        lo, hi = sorted({int(round(ocpu_pct * 100)), int(round(ram_pct * 100))})
        ws_compute["B9"] = f"{lo}–{hi}%" if lo != hi else f"{hi}%"
        if is_ax:
            ws_compute["C9"] = (
                f"% approximation — Ax base OCPU ~{int(round(ocpu_pct*100))}%, RAM ~"
                f"{int(round(ram_pct*100))}%, doubled (×2) for each generation the source "
                "instance is behind the newest OCI generation (1 gen = base, 2 gens = ×2, "
                "3 gens = ×4, …), capped at 95%. Applied to the sizing in columns J and K "
                "as ceil(value × (1 − %)) with a floor of 2, so each row can differ. Price "
                "follows the reduced sizing; this % is not a direct price discount."
            )
        else:
            ws_compute["C9"] = (
                f"% approximation — OCPU reduced ~{int(round(ocpu_pct*100))}%, RAM ~"
                f"{int(round(ram_pct*100))}%, applied to the sizing in columns J and K as "
                "ceil(value × (1 − %)) with a floor of 2. Price follows the reduced "
                "sizing; this % is not a direct price discount."
            )
    else:
        ws_compute["B9"] = "0%"
        ws_compute["C9"] = (
            "% approximation — no compute optimization applied (Rightsize off, or the "
            "selected shape is not eligible). Sizing is carried over as-is."
        )
    # Strip any hidden note/comment or data-validation input prompt the template attached
    # to B9 (it shows as a tooltip on the cell and is separate from the visible C9 text).
    _clear_cell_note(ws_compute, "B9")


MONTH_COLS = "BCDEFGHIJKLM"  # Consumption Ramp grid: months 1..12


def _add_licensing_line(ws_po, windows_monthly):
    """Add a '3rd Party Licensing (Windows)' line to the Pricing Overview baseline and
    re-total. Done by rewriting cells (not insert_rows) so merges / conditional formats /
    images stay intact:  B21 = licensing, B22 = Total Monthly = SUM(B13:B21), B23 = Annual."""
    from copy import copy
    for col in ("A", "B"):
        ws_po[f"{col}23"]._style = copy(ws_po[f"{col}22"]._style)   # annual-total style
        ws_po[f"{col}22"]._style = copy(ws_po[f"{col}21"]._style)   # monthly-total style
        ws_po[f"{col}21"]._style = copy(ws_po[f"{col}20"]._style)   # ordinary line style
    ws_po["A21"] = "3rd Party Licensing (Windows):"
    ws_po["B21"] = round(float(windows_monthly or 0), 2)
    ws_po["A22"] = "Total Monthly Cost:"
    ws_po["B22"] = "=SUM(B13:B21)"
    ws_po["A23"] = "Total Annual Cost:"
    ws_po["B23"] = "=B22*12"


def _add_extra_services(wb, extra_priced):
    """Fold app-added OCI services into the Pricing Overview.

    Each service rolls into the category line it belongs to (Networking -> B18,
    Security -> B19, everything else -> B16), by appending its monthly cost to that line's
    formula. Those lines all sit inside the Total's SUM(B13:B21), so the workbook total
    picks the extras up automatically and still ties out. The added services are also
    itemized on a dedicated sheet for traceability.

    `extra_priced` is the authoritative list from oci_catalog.price_extras().
    """
    import oci_catalog
    if not extra_priced:
        return
    ws = wb["Pricing Overview"]
    sums = {}
    non_storage_into_16 = False
    for s in extra_priced:
        row = oci_catalog.GROUP_TO_OVERVIEW_ROW.get(s["group"], 16)
        sums[row] = sums.get(row, 0.0) + float(s["monthly"] or 0)
        if row == 16 and s["group"] not in ("Storage", "Database"):
            non_storage_into_16 = True

    for row, amount in sums.items():
        cell = ws[f"B{row}"]
        base = cell.value
        add = round(amount, 2)
        if isinstance(base, str) and base.startswith("="):
            cell.value = f"{base}+{add}"
        else:
            cell.value = round(float(base or 0) + add, 2)

    if 16 in sums and non_storage_into_16:
        ws["A16"] = "Storage / Other OCI Services:"

    _itemize_extra_services(wb, extra_priced)


def _itemize_extra_services(wb, extra_priced):
    """Write a dedicated 'Added OCI Services' sheet listing EVERY constituent SKU of each
    app-added service (the estimator's 'Pricing Details'). Pure paper trail — the amounts
    already flow to the Pricing Overview category lines via _add_extra_services, so nothing
    here is summed into any total that feeds the BOM (no double-counting)."""
    if not extra_priced:
        return
    name = "Added OCI Services"
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    for col, w in (("A", 30), ("B", 14), ("C", 42), ("D", 12), ("E", 16),
                   ("F", 14), ("G", 16), ("H", 20)):
        ws.column_dimensions[col].width = w
    ws["A1"] = "Added OCI Services — full SKU breakdown"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    headers = ["Service", "SKU", "SKU Description", "Unit Rate", "Qty / Input",
               "Hours / Month", "Est. Monthly", "Category"]
    hdr_fill = PatternFill("solid", fgColor="FF4472C4")
    for j, h in enumerate(headers, start=1):
        c = ws.cell(3, j, h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
        c.fill = hdr_fill
    # One row per constituent SKU so every service's full SKU list is documented.
    r = 4
    first = r
    for s in extra_priced:
        skus = s.get("skus") or [{"sku": s.get("sku"), "desc": s.get("name"),
                                  "qty": s.get("qty"), "rate": s.get("rate"),
                                  "hours": s.get("hours"), "monthly": s.get("monthly")}]
        for k, sk in enumerate(skus):
            ac = ws.cell(r, 1, s["name"] if k == 0 else "  ↳")
            if k == 0:
                ac.font = Font(name="Calibri", size=11, bold=True)
            ws.cell(r, 2, sk.get("sku"))
            ws.cell(r, 3, sk.get("desc"))
            rc = ws.cell(r, 4, round(float(sk.get("rate") or 0), 4)); rc.number_format = "#,##0.0000"
            ws.cell(r, 5, sk.get("qty"))
            hrs = sk.get("hours")
            ws.cell(r, 6, hrs if isinstance(hrs, (int, float)) and hrs else None)
            mc = ws.cell(r, 7, round(float(sk.get("monthly") or 0), 2)); mc.number_format = "#,##0.00"
            ws.cell(r, 8, s.get("group"))
            r += 1
    tc = ws.cell(r, 1, "Total Added OCI Services"); tc.font = Font(name="Calibri", size=12, bold=True)
    tot = ws.cell(r, 7, f"=SUM(G{first}:G{r - 1})")
    tot.number_format = "#,##0.00"; tot.font = Font(name="Calibri", size=12, bold=True)
    ws.freeze_panes = "A4"


def _repoint_ramp_refs(ws_po, months, include_windows=False):
    """The cost-profile's Year-1 columns sum the ramp's consumption %. Re-point those
    ranges at the actual number of ramp months the app is set to (the template hardcodes
    F12:F23 = 12 months). Year 1 uses at most the first 12 ramp months."""
    y1_last = RAMP_FIRST_MONTH_ROW + min(12, max(1, months)) - 1
    rng = f"'Consumption Ramp'!$F${RAMP_FIRST_MONTH_ROW}:$F${y1_last}"
    # Core infrastructure = compute; Windows 3rd-party licensing (B21) is folded in and
    # ramped alongside it only when it's present in the BOM.
    core = "SUM($B$13:$B$15,$B$21)" if include_windows else "SUM($B$13:$B$15)"
    ws_po["E14"] = f"=IFERROR({core}*SUM({rng}),0)"
    ws_po["F14"] = f"=IFERROR({core}*12,0)"
    ws_po["E15"] = f"=IFERROR($B$16*SUM({rng}),0)"
    ws_po["E17"] = f"=IFERROR($B$20*SUM({rng}),0)"


def _zero_unmodeled_sheets(wb):
    """The app doesn't model Networking, Security/KMS, DR or Storage-Backups yet, so the
    reference deliverable's hardcoded numbers are cleared — otherwise they'd inject cost
    the app never shows. Sheet structure + architecture notes are preserved."""
    net = wb["Networking"]
    for r in range(11, 19):            # component rows -> Networking!G19 total becomes 0
        for c in range(1, 9):
            net.cell(r, c).value = None
    kms = wb["Security KMS"]
    for r in range(11, 13):            # -> 'Security KMS'!B5 becomes 0
        for c in range(1, 10):
            kms.cell(r, c).value = None
    dr = wb["DR"]
    for r in list(range(12, 16)) + list(range(19, 23)) + list(range(26, 30)):
        dr.cell(r, 2).value = "No"     # "Include in Price" -> DR!B5 becomes 0
    anx = wb["Annexure Addendum to Storage"]
    anx["D6"] = None                   # drop the reference capacity input
    anx["J6"] = "No"
    anx["J7"] = "No"                   # -> Pricing Overview!B17 (SUMIF) becomes 0


# Sheets that always stay visible even when "empty" — they carry the deliverable's
# framing, rates, totals and ramp, so they're meaningful regardless of inventory.
_CORE_SHEETS = {
    "Table of Contents", "Assumptions", "Rate Card", "Pricing Overview",
    "Compute", "Consumption Ramp",
}


def _sheet_has_values(ws, rows, cols):
    """True if any cell in the given rows/cols holds a non-blank, non-zero value."""
    for r in rows:
        for c in cols:
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s and s not in ("0", "0.0", "$0", "$0.00", "-"):
                return True
    return False


def _shape_used_label(shape_used):
    """A readable shape name from a row's shapeUsed (dict payload or plain string)."""
    if isinstance(shape_used, dict):
        return _clean(shape_used.get("shortLabel") or shape_used.get("label")
                      or shape_used.get("key"))
    return _clean(shape_used)


def _place_after(wb, name, after):
    """Move sheet `name` to sit immediately after sheet `after` (no-op if either is missing)."""
    if name not in wb.sheetnames or after not in wb.sheetnames or name == after:
        return
    s = wb[name]
    wb._sheets.remove(s)
    idx = wb.sheetnames.index(after) + 1
    wb._sheets.insert(idx, s)


def _hide_empty_sheets(wb, apps, storage_rows):
    """Hide printout sheets that ended up with no data so the deliverable doesn't ship
    empty sections. Core sheets always stay visible; the rest are hidden (not
    veryHidden, so a user can unhide) when their data region is blank. Hidden sheets
    still compute, so the Pricing Overview references into them are unaffected."""
    def hide(name):
        if name in wb.sheetnames and name not in _CORE_SHEETS:
            wb[name].sheet_state = "hidden"

    if not apps:
        hide("Applications Migrated to OCI")
    if not storage_rows:
        hide("Storage")
    if "Networking" in wb.sheetnames and not _sheet_has_values(
            wb["Networking"], range(11, 19), range(1, 9)):
        hide("Networking")
    if "Security KMS" in wb.sheetnames and not _sheet_has_values(
            wb["Security KMS"], range(11, 13), range(1, 10)):
        hide("Security KMS")
    if "DR" in wb.sheetnames:
        dr = wb["DR"]
        included = any(str(dr.cell(r, 2).value).strip().lower() == "yes"
                       for r in list(range(12, 16)) + list(range(19, 23)) + list(range(26, 30)))
        if not included:
            hide("DR")
    if "Annexure Addendum to Storage" in wb.sheetnames:
        if wb["Annexure Addendum to Storage"]["D6"].value in (None, "", 0):
            hide("Annexure Addendum to Storage")
    # Never leave the active sheet hidden.
    if wb.active is not None and wb.active.sheet_state != "visible":
        for ws in wb.worksheets:
            if ws.sheet_state == "visible":
                wb.active = wb.index(ws)
                break


RAMP_FIRST_MONTH_ROW = 12
RAMP_TPL_LAST_MONTH_ROW = 23      # the reference deliverable ships 12 months
RAMP_TPL_GRID_TITLE_ROW = 27      # "Detailed Consumption Ramp" block starts here
RAMP_TPL_GRID_LAST_ROW = 38
RAMP_MAX_MONTHS = 60

# Component grid rows (relative to the template) -> the Pricing Overview line they scale.
# The consumption ramp scales the OCI SERVICES only (B13:B20), matching the app's ramp,
# whose ceiling is the OCI-services monthly total (pricing.totals.monthly) and excludes
# Windows 3rd-party licensing. Windows (B21) is a separate flat license line, not ramped.
_GRID_COMPONENTS = [
    ("Compute (OCPUs)", "$B$13"),
    ("RAM", "$B$14"),
    ("VM Block Storage", "$B$15"),
    ("Storage (Object/File)", "$B$16"),
    ("Storage Backups", "$B$17"),
    ("Networking + Security / KMS", "$B$18"),
    ("Disaster Recovery", "$B$20"),
]


def _ramp_percentages(ramp):
    """The app's ramp as monthly consumption fractions (month spend / steady state).
    Length follows the app's ramp-months toggle."""
    if not isinstance(ramp, dict):
        return None
    ceiling = float(ramp.get("ceiling") or 0)
    monthly = [float(x) for x in (ramp.get("monthly") or [])]
    if ceiling <= 0 or not monthly:
        return None
    return [max(0.0, min(1.0, m / ceiling)) for m in monthly[:RAMP_MAX_MONTHS]]


def _populate_ramp(ws, ramp, include_windows=False):
    """Rebuild the Consumption Ramp for EXACTLY the number of months the app's ramp
    toggle is set to, driven by the app's curve. When Windows 3rd-party licensing is
    present (include_windows), it's added as a ramped component so both ramps carry it.

    Returns the month count so the Pricing Overview's F-range refs can be re-pointed.
    """
    from copy import copy
    from openpyxl.utils import get_column_letter

    # Windows licensing is ramped as its own component only when it's actually in the BOM.
    components = list(_GRID_COMPONENTS)
    if include_windows:
        components.append(("3rd Party Licensing", "$B$21"))

    pcts = _ramp_percentages(ramp)
    n = len(pcts) if pcts else 12
    n = max(1, min(n, RAMP_MAX_MONTHS))

    # The area carries merges; drop them so every cell is writable while we rebuild.
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= RAMP_FIRST_MONTH_ROW:
            ws.unmerge_cells(str(rng))

    def sty(r, c):
        return copy(ws.cell(r, c)._style)

    # ---- capture prototypes from the reference layout BEFORE clearing ----
    month_sty = {c: sty(12, c) for c in range(1, 9)}
    wave_text = [[ws.cell(r, c).value for c in range(2, 6)] for r in range(12, 24)]
    grid_sty = {r: {c: sty(r, c) for c in (1, 2)} for r in range(27, 39)}
    grid_title = ws.cell(27, 1).value
    wave_block = [([ws.cell(r, c).value for c in range(1, 8)],
                   {c: sty(r, c) for c in range(1, 8)}) for r in range(40, 47)]
    chart_hdr = ([ws.cell(50, c).value for c in (1, 2)], {c: sty(50, c) for c in (1, 2)})
    chart_sty = {c: sty(51, c) for c in (1, 2)}

    # ---- clear everything below the month header ----
    for r in range(RAMP_FIRST_MONTH_ROW, 12 + RAMP_MAX_MONTHS + 60):
        for c in range(1, 2 + RAMP_MAX_MONTHS + 2):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # ---- layout: months -> grid -> wave sequencing -> chart source ----
    m_first, m_last = RAMP_FIRST_MONTH_ROW, RAMP_FIRST_MONTH_ROW + n - 1
    g_title = m_last + 4
    g_head, g_pct = g_title + 1, g_title + 2
    g_first = g_pct + 1
    g_last = g_first + len(components) - 1
    g_cum = g_last + 1
    w_title = g_cum + 3
    c_head = w_title + len(wave_block) + 2
    total_col = 2 + n

    # months
    for i in range(n):
        r = m_first + i
        cl = get_column_letter(2 + i)
        for c in range(1, 9):
            ws.cell(r, c)._style = copy(month_sty[c])
        ws.cell(r, 1).value = i + 1
        for j, v in enumerate(wave_text[min(i, len(wave_text) - 1)]):
            ws.cell(r, 2 + j).value = v
        ws.cell(r, 6).value = round(pcts[i], 6) if pcts else None
        ws.cell(r, 7).value = f"=SUM({cl}${g_first}:{cl}${g_last})"
        ws.cell(r, 8).value = f"=SUM($G${m_first}:G{r})"

    # component grid
    ws.cell(g_title, 1).value = grid_title
    ws.cell(g_title, 1)._style = copy(grid_sty[27][1])
    ws.cell(g_head, 1).value = "Component"
    ws.cell(g_head, 1)._style = copy(grid_sty[28][1])
    ws.cell(g_pct, 1).value = "Consumption %"
    ws.cell(g_pct, 1)._style = copy(grid_sty[29][1])
    for i in range(n):
        c = 2 + i
        ws.cell(g_head, c).value = f"Month {i + 1}"
        ws.cell(g_head, c)._style = copy(grid_sty[28][2])
        ws.cell(g_pct, c).value = f"=F{m_first + i}"
        ws.cell(g_pct, c)._style = copy(grid_sty[29][2])
    ws.cell(g_head, total_col).value = "Total"
    ws.cell(g_head, total_col)._style = copy(grid_sty[28][2])

    last_month_col = get_column_letter(1 + n)
    for k, (label, po_ref) in enumerate(components):
        r = g_first + k
        src = grid_sty[min(30 + k, 36)]
        ws.cell(r, 1).value = label
        ws.cell(r, 1)._style = copy(src[1])
        for i in range(n):
            c = 2 + i
            cl = get_column_letter(c)
            ws.cell(r, c).value = (
                f"=('Pricing Overview'!$B$18+'Pricing Overview'!$B$19)*{cl}{g_pct}"
                if po_ref == "$B$18" else f"='Pricing Overview'!{po_ref}*{cl}{g_pct}")
            ws.cell(r, c)._style = copy(src[2])
        ws.cell(r, total_col).value = f"=SUM(B{r}:{last_month_col}{r})"
        ws.cell(r, total_col)._style = copy(src[2])

    ws.cell(g_cum, 1).value = "Cumulative Total"
    ws.cell(g_cum, 1)._style = copy(grid_sty[37][1])
    for i in range(n):
        c = 2 + i
        cl = get_column_letter(c)
        ws.cell(g_cum, c).value = f"=SUM($B${g_first}:{cl}{g_last})"
        ws.cell(g_cum, c)._style = copy(grid_sty[37][2])
    ws.cell(g_cum, total_col).value = f"=SUM(B{g_cum}:{last_month_col}{g_cum})"
    ws.cell(g_cum, total_col)._style = copy(grid_sty[37][2])

    # wave sequencing reference table (static)
    for j, (vals, styles) in enumerate(wave_block):
        r = w_title + j
        for c in range(1, 8):
            ws.cell(r, c).value = vals[c - 1]
            ws.cell(r, c)._style = copy(styles[c])

    # chart-source table: Month / Cumulative Total, re-pointed at the new cumulative row
    for c in (1, 2):
        ws.cell(c_head, c).value = chart_hdr[0][c - 1]
        ws.cell(c_head, c)._style = copy(chart_hdr[1][c])
    for i in range(n):
        r = c_head + 1 + i
        cl = get_column_letter(2 + i)
        ws.cell(r, 1).value = f"Month {i + 1}"
        ws.cell(r, 1)._style = copy(chart_sty[1])
        ws.cell(r, 2).value = f"={cl}{g_cum}"
        ws.cell(r, 2)._style = copy(chart_sty[2])

    ws["A10"] = f"{n}-Month Consumption Ramp"
    ws["B5"] = "='Pricing Overview'!$B$22"
    return n


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def build_full_bom_bytes(pricing, rows=None, fields=None, ramp=None, bom_name="",
                         shape=None, hours=None, block_rate=None, vpu_rate=None,
                         default_vpus=None, file_rate=None, windows_rate=None,
                         windows_sku=None, optimization=0.0, include_diagram=True,
                         extra_services=None, cloud_comparison=None):
    """Build the 12-sheet Full BOM workbook populated from the app's priced inventory.

    The workbook is built to tie out to the app EXACTLY:
      Compute (OCPU+RAM+block) + Storage + 3rd-party licensing = the app's monthly total,
      and the Consumption Ramp follows the app's ramp curve. Sheets the app does not model
      yet (Networking / Security KMS / DR / Backups) are zeroed so they inject no cost.

    EVERY rate is supplied by the caller from the app's catalog. These deliberately have NO
    numeric defaults: a default here is a second copy of the price list, and a stale copy is
    exactly how a deliverable ends up quoting last year's rates. If the caller doesn't pass
    a rate, we pull it from app.py rather than from whatever the source template shipped.
    """
    if None in (hours, block_rate, vpu_rate, default_vpus, file_rate, windows_rate,
                windows_sku):
        import app as _app                       # the single source of truth for pricing
        hours = _app.HOURS_PER_MONTH if hours is None else hours
        block_rate = _app.storage_rate("B91961") if block_rate is None else block_rate
        vpu_rate = _app.storage_rate("B91962") if vpu_rate is None else vpu_rate
        file_rate = _app.storage_rate("B89057") if file_rate is None else file_rate
        default_vpus = (_app.BLOCK_PERFORMANCE_UNITS_PER_GB
                        if default_vpus is None else default_vpus)
        windows_rate = _app.WINDOWS_LICENSE_RATE if windows_rate is None else windows_rate
        windows_sku = _app.WINDOWS_LICENSE_SKU if windows_sku is None else windows_sku
    spec = load_spec()
    wb = build_workbook(spec)

    keys = _resolve_inventory_keys(fields or [])
    raw_by_id = {}
    for r in (rows or []):
        rid = r.get("__id") or r.get("rowId")
        if rid:
            raw_by_id[str(rid)] = r

    servers = []
    apps = []
    storage_rows = []
    is_cloud_bill = bool(cloud_comparison)
    for pr in (pricing or {}).get("rows", []):
        specs = pr.get("specs") or {}
        raw = raw_by_id.get(str(pr.get("rowId"))) or {}

        # Cloud-bill mode: only the Compute-category bill lines belong on the Compute sheet.
        # Storage/networking/etc. service rows roll into the Pricing Overview lines instead,
        # so including them here would double-count them.
        if is_cloud_bill and (pr.get("ociServiceCategory") or "") != "Compute":
            continue

        def rv(role):
            k = keys.get(role)
            return _clean(raw.get(k)) if k else ""

        vcpu = specs.get("vcpus") or (float(specs.get("ocpus") or 0) * 2)
        virt = rv("virt")
        # Normalize to the template's Virtual/Physical vocabulary (drives the OCPU formula).
        vl = _norm(virt)
        if "phys" in vl:
            virt = "Physical"
        elif vl:
            virt = "Virtual"

        # Only use a REAL application column. If the inventory has none, leave Master
        # Application blank (and the Applications sheet empty) rather than duplicating
        # the server name — no invented data. Cloud bills have no application grouping,
        # so never carry an "app" here (the inventory-key match can otherwise land on an
        # unrelated column like mapping confidence).
        app_name = "" if is_cloud_bill else rv("app")
        srv = {
            "server": rv("server") or _clean(pr.get("name")),
            "tier": rv("tier"),
            "env": rv("env") or _clean(pr.get("environment")),
            "app": app_name,
            "desc": rv("desc"),
            "virt": virt,
            "os_name": rv("os_name"),
            "os_family": rv("os_family"),
            # Full precision — rounding here would put the workbook a few cents off the
            # app's total across hundreds of rows, and the two must tie out exactly.
            "vcpu": float(vcpu) if vcpu else None,
            "memory_gb": float(specs.get("memoryGb") or 0) or None,
            "storage_gb": float(specs.get("blockStorageGb") or 0) or None,
            # The OCI shape this server maps to (shown on the Compute sheet). shapeUsed can
            # be a dict (on-prem shape payload) or a string (cloud-bill shape name).
            "shape": _shape_used_label(pr.get("shapeUsed")),
            # Per-row monthly hours from the data source (the app already priced each row at
            # its own hours). Falls back to the global hours only when the row has none.
            # Cloud-bill: use the EFFECTIVE hours implied by the bill's metered usage
            # (OCPU-hours / OCPU) so OCPU x hours x rate reproduces the app's actual cost —
            # a bill line can cover far more than one instance's 730 hours.
            "hours": (_cloud_effective_hours(pr) if is_cloud_bill
                      else float(pr.get("hoursPerMonth") or 0) or None),
        }
        if not (srv["vcpu"] or srv["memory_gb"] or srv["storage_gb"]):
            continue
        servers.append(srv)
        if app_name and app_name not in apps:
            apps.append(app_name)

        fs = float(specs.get("fileStorageGb") or 0)
        if fs > 0:
            storage_rows.append({
                "server": srv["server"], "tier": srv["tier"], "env": srv["env"],
                "app": app_name, "signal": "Shared / file storage",
                "target": "OCI File Storage", "gb": round(fs, 2), "rate": file_rate,
            })

    # Cloud-bill: itemize the storage services on the Storage sheet (the compute-loop above
    # skipped them). This is display only — the Storage line total is set on the Pricing
    # Overview by _add_cloud_bill_services, so there's no double count.
    if cloud_comparison:
        storage_rows = _cloud_storage_rows(cloud_comparison.get("pricing") or pricing)
    windows_monthly = sum(float(r.get("windowsLicenseMonthly") or 0)
                          for r in (pricing or {}).get("rows", []))
    # Build the Rate Card FIRST — from only the SKUs/rates used in this build, sorted
    # alphabetically — so the Compute/Storage formulas can reference the exact cells it
    # placed each rate on (transparency, and it ties out to the app).
    shape_label = (shape or {}).get("shortLabel") or (shape or {}).get("label") or ""
    rate_entries = _collect_rate_card_entries(
        shape, block_rate, vpu_rate, default_vpus, hours, file_rate, windows_rate,
        windows_sku, windows_monthly > 0, servers, storage_rows, pricing,
        extra_services, is_cloud_bill)
    rate_refs = _write_rate_card(wb["Rate Card"], rate_entries)
    _populate_compute(wb[COMPUTE_SHEET], servers, hours, rate_refs, shape_label)
    _populate_apps(wb[APPS_SHEET], apps)
    _populate_storage(wb[STORAGE_SHEET], storage_rows, rate_refs, file_rate)
    _set_toc(wb["Table of Contents"], bom_name)
    _apply_customer_name(wb, bom_name)
    # Compute optimization: record the % the app's Rightsize applied. Ax = 15% OCPU /
    # 20% RAM, regular E6 = 10% / 15%; the app already shrank the quantities to match.
    rightsized = any(r.get("rightsized") for r in (pricing or {}).get("rows", []))
    shape_key = str((shape or {}).get("key") or "")
    is_ax = shape_key.endswith("-ax")
    if is_ax:
        ocpu_pct, ram_pct = 0.15, 0.20
    elif shape_key == "e6-standard":
        ocpu_pct, ram_pct = 0.10, 0.15
    else:
        ocpu_pct = ram_pct = 0.0
    _set_optimization(wb[COMPUTE_SHEET], rightsized, ocpu_pct, ram_pct, is_ax)
    _zero_unmodeled_sheets(wb)
    _add_licensing_line(wb["Pricing Overview"], windows_monthly)
    # App-added OCI services roll into the matching Pricing Overview lines (which sit inside
    # the total) and are itemized on the Networking sheet.
    if extra_services:
        import oci_catalog
        priced, _ = oci_catalog.price_extras(extra_services, hours)
        _add_extra_services(wb, priced)
    # Cloud-bill mode: roll the non-compute mapped services into the Pricing Overview lines
    # so the template total covers the whole bill, not just compute.
    if cloud_comparison:
        _add_cloud_bill_services(wb, cloud_comparison.get("pricing") or pricing)
        # Summarize all 11 product-group topics on the Pricing Overview and add a detail
        # sheet for each group that has cost and lacks a dedicated sheet.
        _add_product_group_topics(wb, cloud_comparison.get("pricing") or pricing)
    ramp_months = _populate_ramp(wb["Consumption Ramp"], ramp, include_windows=windows_monthly > 0)
    _repoint_ramp_refs(wb["Pricing Overview"], ramp_months, include_windows=windows_monthly > 0)

    # Architecture diagram generated from THIS BOM (deterministic; no model call).
    # If the diagram toolchain isn't available the export still succeeds — the template's
    # reference picture just stays in place.
    arch_png = None
    if include_diagram:
        try:
            import bom_diagram
            _, arch_png = bom_diagram.build_architecture(
                pricing, rows, keys, bom_name,
                (shape or {}).get("shortLabel") or (shape or {}).get("label") or "",
                sites=_distinct_sites(fields or [], rows or []))
            if arch_png:
                embed_architecture(wb["Pricing Overview"], arch_png,
                                   spec.get("architecture_anchor"))
        except Exception:
            # Don't swallow the reason silently — a missing diagram was undebuggable.
            import traceback
            traceback.print_exc()
            arch_png = None
    if not arch_png:
        # Nothing to show: the template ships no architecture picture of its own, so the
        # slot is simply empty. Clear its caption too.
        ws_po = wb["Pricing Overview"]
        if isinstance(ws_po["D27"].value, str):
            ws_po["D27"] = None

    # Cloud-bill mode: append the AWS->OCI bill sheets (Product Breakdown, Service Mapping,
    # Notes, Cloud Bill Overview) so every mapped service and its mapping is preserved
    # alongside the 12-sheet deliverable — nothing from the bill printout is lost.
    if cloud_comparison:
        try:
            import bom_export
            bom_export.add_cloud_comparison_sheets(
                wb, cloud_comparison.get("pricing") or {"rows": rows, "totals": {}},
                cloud_comparison.get("ramp"), bom_name,
                cloud_comparison.get("ociDiscount") or 0.0,
                cloud_comparison.get("extraServices"),
                cloud_comparison.get("hours") or hours, use_active=False)
        except Exception:
            import traceback
            traceback.print_exc()
        # Service Mapping reads right after the Pricing Overview.
        _place_after(wb, "Service Mapping", "Pricing Overview")

    # Hide any printout sheet that ended up with no data (empty Storage/Networking/DR/
    # Security KMS/Applications/Annexure), so the deliverable doesn't ship blank sections.
    _hide_empty_sheets(wb, apps, storage_rows)

    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tmp_path = tf.name
    wb.save(tmp_path)
    _postprocess(tmp_path, spec)
    data = Path(tmp_path).read_bytes()
    Path(tmp_path).unlink(missing_ok=True)
    return data
