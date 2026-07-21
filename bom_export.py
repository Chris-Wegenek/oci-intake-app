"""Excel BOM exporter.

Reproduces the two sheets created by the "E6 Ax BOM Creator" Office Script
("BOM w E6 Acceleron" + "Overview") using openpyxl, so the app can export a
formatted workbook that matches the script's layout, pricing, and formulas.

The pricing rates here intentionally mirror the BOM script exactly:
  - E6 Acceleron OCPU      B112530   $0.0138 / OCPU-hour  x 730
  - E6 Acceleron Memory    B112530   $0.0108 / GB-hour    x 730
  - Block Volume Storage   B91961    $0.0255 / GB-month
  - Block Volume Perf Unit B91962    $0.0017 / unit-month (10 units per GB)
  - Windows OS license     B88318    $0.0920 / OCPU-hour  x 730
  - FastConnect 1 Gbps     B88326    $0.2125 / port-hour  x 730
"""

import json
import os
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---- Pricing constants (must match the BOM script) ----
HOURS = 730
CPU_RATE = 0.0138
MEM_RATE = 0.0108
DISK_RATE = 0.0255
PERF_RATE = 0.0017
PERF_UNITS_PER_GB = 10
WINDOWS_RATE = 0.0920
FASTCONNECT_RATE = 0.2125

# ---- Colors (hex from the script, ARGB for openpyxl) ----
HDR_FILL = "C0E6F5"
NAME_RED = "FF9999"
BV_PINK = "FFE8E8"
BORDER = "000000"
OV_NAVY = "7F1D1D"
OV_BLUE = "C0504D"
OV_LIGHT = "F2DCDB"
OV_INPUT = "FFF2CC"
OV_INPUT_BORDER = "BF8F00"

MONEY2 = '"$"#,##0.00'
MONEY4 = '"$"#,##0.0000'
MONEY0 = '"$"#,##0'
PCT0 = "0%"

# Default shape (E6 Acceleron) used when the caller does not pass shape details.
DEFAULT_SHAPE = {
    "label": "E6 Acceleron",
    "shortLabel": "E6 Acceleron",
    "computeSku": "B112530",
    "memorySku": "B112530",
    "computeRate": CPU_RATE,
    "memoryRate": MEM_RATE,
}


def _resolve_shape(shape):
    merged = dict(DEFAULT_SHAPE)
    if shape:
        for key in ("label", "shortLabel", "computeSku", "memorySku", "computeRate", "memoryRate"):
            if shape.get(key) not in (None, ""):
                merged[key] = shape[key]
    return merged


def _fill(hexcolor):
    return PatternFill("solid", fgColor="FF" + hexcolor)


def _side(color=BORDER, style="medium"):
    return Side(style=style, color="FF" + color)


def _box(ws, r1, c1, r2, c2, color=BORDER, style="medium"):
    """Apply an outer border around a rectangular range."""
    side = _side(color, style)
    for c in range(c1, c2 + 1):
        top = ws.cell(row=r1, column=c)
        bot = ws.cell(row=r2, column=c)
        top.border = Border(top=side, left=top.border.left, right=top.border.right, bottom=top.border.bottom)
        bot.border = Border(bottom=side, left=bot.border.left, right=bot.border.right, top=bot.border.top)
    for r in range(r1, r2 + 1):
        left = ws.cell(row=r, column=c1)
        right = ws.cell(row=r, column=c2)
        left.border = Border(left=side, top=left.border.top, right=left.border.right, bottom=left.border.bottom)
        right.border = Border(right=side, top=right.border.top, left=right.border.left, bottom=right.border.bottom)


CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left")


def _detect_os(server):
    text = " ".join(str(server.get(k, "")) for k in ("os", "name", "environment", "raw")).lower()
    if "windows" in text:
        return "windows"
    if "linux" in text:
        return "linux"
    return ""


def build_bom_sheet(ws, servers, shape=None, hide_windows=False, hours=HOURS):
    """Render the BOM sheet for the selected shape. Returns the windows OCPU total."""
    shape = _resolve_shape(shape)
    headers = ["Part", "Description", "Part Qty", "Instance Qty", "Usage Qty (Hours)",
               "Unit Price", "Monthly Cost", "VM Cost"]
    for i, text in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=text)
        c.fill = _fill(HDR_FILL)
        c.font = Font(bold=True, size=11)
        c.alignment = CENTER if i in (5, 7, 8) else Alignment(horizontal="left", vertical="center")
    ws.cell(row=3, column=9).fill = _fill(HDR_FILL)

    # --- Free Tier block ---
    ws.cell(row=4, column=2, value="Free Tier").font = Font(bold=True, size=17)
    free_rows = [
        "OCI Data Ingress (Inbound Data Transfer)",
        "OCI Data Egress (Outbound) (first 10TB)",
        "Flexible Network Load Balancer",
        "OCI Vault Secrets Security",
        "OCI User Management and MFA",
        "OCI Cloud Guard (Monitoring and Defense)",
        "OCI Data Encryption",
        "OCI Network Firewall",
        "Support",
    ]
    for i, label in enumerate(free_rows):
        r = 5 + i
        ws.cell(row=r, column=2, value=label).font = Font(bold=True, size=11)
        ws.cell(row=r, column=3, value=("Included" if label == "Support" else "Free")).font = Font(size=11)
        g = ws.cell(row=r, column=7, value=0)
        g.font = Font(bold=True)
        g.number_format = MONEY2

    # --- FastConnect ---
    ws.cell(row=16, column=2, value="FastConnect").font = Font(bold=True, size=17)
    _line(ws, 17, "B88326", "          OCI - FastConnect 1 Gbps (Port Hour)", 0, 1, hours, FASTCONNECT_RATE)

    # --- Windows Licenses (qty filled after blocks) ---
    ws.cell(row=19, column=2, value="Windows Licenses").font = Font(bold=True, size=17)
    _line(ws, 20, "B88318", "          Compute - Windows OS (OCPU Per Hour)", 0, 1, hours, WINDOWS_RATE)

    # --- Virtual Machines header ---
    ws.cell(row=22, column=2, value="Virtual Machines").font = Font(bold=True, size=17)

    block_size = 7
    first_name_row = 24  # Excel row of the first server name
    windows_cpu_total = 0

    for b, server in enumerate(servers):
        name_row = first_name_row + b * block_size
        cpu_row, ram_row, bv_row, disk_row, perf_row = (name_row + k for k in range(1, 6))

        ocpus = int(server.get("ocpus") or 0)
        memory = int(server.get("memory") or 0)
        disk = float(server.get("disk") or 0)
        if _detect_os(server) == "windows" and not hide_windows:
            windows_cpu_total += ocpus

        # Server name row: fill across A-H. Color by feasibility against the OCI shape.
        status = server.get("sizeStatus", "ok")
        if status == "impossible":
            row_fill, name_color = _fill("C00000"), "FFFFFFFF"   # strong red = cannot be built
        elif status == "baremetal":
            row_fill, name_color = _fill("FFC000"), "FF000000"   # orange = needs bare metal
        else:
            row_fill, name_color = _fill(NAME_RED), "FF000000"
        for c in range(1, 9):
            ws.cell(row=name_row, column=c).fill = row_fill
        nm = ws.cell(row=name_row, column=2, value=server.get("name", f"Server {b + 1}"))
        nm.font = Font(bold=True, size=15, color=name_color)
        nm.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
        if status != "ok" and server.get("sizeMessage"):
            note = ws.cell(row=name_row, column=10,
                           value=("IMPOSSIBLE: " if status == "impossible" else "NEEDS BARE METAL: ") + server["sizeMessage"])
            note.font = Font(bold=True, size=10, color=("FFC00000" if status == "impossible" else "FFBF8F00"))

        srv_hours = server.get("hours") or hours
        # Use this server's own shape (Auto/processor-matching) when present,
        # otherwise the workbook's selected shape.
        srv_shape = _resolve_shape(server.get("shape")) if server.get("shape") else shape
        _line(ws, cpu_row, srv_shape["computeSku"],
              f" Compute - {srv_shape['label']} - OCPU (OCPU Per Hour)      Capacity Type: On - Demand",
              ocpus, 1, srv_hours, srv_shape["computeRate"])
        _line(ws, ram_row, srv_shape["memorySku"],
              f" Compute - {srv_shape['label']} - Memory(Gigabyte Per Hour)    Capacity Type: On - Demand",
              memory, 1, srv_hours, srv_shape["memoryRate"])

        bv = ws.cell(row=bv_row, column=2, value=" Boot Volume (Local Storage Sizes)")
        bv.fill = _fill(BV_PINK)
        bv.font = Font(bold=True)

        _line(ws, disk_row, "B91961",
              "Storage - Block Volume - Storage (Gigabyte Storage Capacity Per Month)",
              disk, 1, 1, DISK_RATE)
        _line(ws, perf_row, "B91962",
              "Storage - Block Volume - Performance Units (Performance Units Per Gigabyte Per Month)",
              PERF_UNITS_PER_GB * disk, 1, 1, PERF_RATE)

        # Col H VM cost box
        h = ws.cell(row=disk_row, column=8, value="Total VM Cost")
        h.font = Font(bold=True)
        h.alignment = CENTER
        tot = ws.cell(row=perf_row, column=8, value=f"=SUM(G{cpu_row}:G{perf_row})")
        tot.font = Font(bold=True)
        tot.alignment = CENTER
        tot.number_format = MONEY2

        _box(ws, name_row, 1, perf_row, 8, BORDER, "medium")
        _box(ws, disk_row, 8, perf_row, 8, BORDER, "medium")

    # Windows license quantity -> C20, drives G20
    if servers:
        ws.cell(row=20, column=3, value=windows_cpu_total).alignment = CENTER

    # --- Totals ---
    total_row = first_name_row + len(servers) * block_size
    ws.cell(row=total_row, column=7, value="Total Per Month:").font = Font(bold=True, size=15)
    g_tot = ws.cell(row=total_row, column=8, value="=SUM(G:G)")
    g_tot.font = Font(bold=True, size=15)
    g_tot.number_format = MONEY2

    # Summary cells in cols I/J referenced by the Overview sheet
    ws.cell(row=14, column=9, value="Total Per Month:").font = Font(bold=True, size=15)
    j15 = ws.cell(row=15, column=10, value="=SUM(G:G)")
    j15.font = Font(bold=True, size=15)
    j15.number_format = MONEY2
    ws.cell(row=17, column=9, value="Total Per Year:").font = Font(bold=True, size=15)
    j18 = ws.cell(row=18, column=10, value="=J15*12")
    j18.font = Font(bold=True, size=15)
    j18.number_format = MONEY0

    # Column number formats / widths
    for r in range(1, total_row + 2):
        ws.cell(row=r, column=6).number_format = MONEY4
        ws.cell(row=r, column=7).number_format = MONEY2
        ws.cell(row=r, column=8).number_format = MONEY2
    widths = {1: 12, 2: 62, 3: 10, 4: 12, 5: 16, 6: 12, 7: 14, 8: 16, 9: 18, 10: 16}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    return windows_cpu_total


def _line(ws, row, part, desc, qty, instance_qty, hours, unit_price):
    """Write a standard BOM line row (cols A-G) with the G=C*D*E*F formula."""
    ws.cell(row=row, column=1, value=part).font = Font(bold=True)
    ws.cell(row=row, column=2, value=desc).font = Font(size=11)
    ws.cell(row=row, column=3, value=qty).alignment = CENTER
    ws.cell(row=row, column=4, value=instance_qty).alignment = CENTER
    ws.cell(row=row, column=5, value=hours).alignment = CENTER
    f = ws.cell(row=row, column=6, value=unit_price)
    f.alignment = CENTER
    f.number_format = MONEY4
    g = ws.cell(row=row, column=7, value=f"=C{row}*D{row}*E{row}*F{row}")
    g.number_format = MONEY2


def _build_extra_services_sheet(wb, priced, total):
    """Itemize app-added OCI services on their own sheet in the Quick BOM."""
    ws = wb.create_sheet("Additional Services")
    ws.column_dimensions["A"].width = 34
    for col in "BCDE":
        ws.column_dimensions[col].width = 16
    ws["A1"] = "Additional OCI Services (configured in app)"
    ws["A1"].font = Font(bold=True, size=14)
    for j, h in enumerate(["Service", "Category", "SKU", "Sizing", "Monthly (USD)"], start=1):
        c = ws.cell(2, j, h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = _fill(OV_BLUE)
    for i, s in enumerate(priced):
        r = 3 + i
        ws.cell(r, 1, s["name"])
        ws.cell(r, 2, s["group"])
        ws.cell(r, 3, s["sku"])
        ws.cell(r, 4, s["sizing"])
        ws.cell(r, 5, round(float(s["monthly"] or 0), 2)).number_format = MONEY2
    tr = 3 + len(priced)
    ws.cell(tr, 4, "Total Monthly:").font = Font(bold=True)
    tc = ws.cell(tr, 5, round(float(total or 0), 2))
    tc.font = Font(bold=True)
    tc.number_format = MONEY2


def build_overview_sheet(ws, util_by_year, existing_infra_cost, bom_sheet_name="BOM w E6 Acceleron", existing_label="Existing Infra Cost (enter):", oci_discount=0.0, extra_oci=0.0, extra_third_party=0.0):
    """Render the 'Overview' sheet, pulling the 5-year utilization ramp from the app.
    oci_discount (0-1) is applied to the OCI totals and shown in a discount cell."""
    BOM = f"'{bom_sheet_name}'"
    ws.column_dimensions["A"].width = 28
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 16

    def band(cell, text, size=13):
        c = ws[cell]
        c.value = text
        c.fill = _fill(OV_NAVY)
        c.font = Font(bold=True, color="FFFFFFFF", size=size)
        c.alignment = CENTER

    def colhead(cell, text):
        c = ws[cell]
        c.value = text
        c.fill = _fill(OV_BLUE)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.alignment = CENTER

    # Title banner
    ws.merge_cells("A1:H2")
    t = ws["A1"]
    t.value = "Cloud Migration Cost Overview"
    t.fill = _fill(OV_NAVY)
    t.font = Font(bold=True, color="FFFFFFFF", size=26, name="Calibri")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26

    # OCI discount applied from the app (shown + applied to the OCI totals).
    band("G4", "OCI Discount*")
    ws["G4"].fill = _fill(OV_BLUE)
    gd = ws["G5"]
    gd.value = round(float(oci_discount or 0), 4)
    gd.fill = _fill(OV_INPUT)
    gd.font = Font(bold=True, size=18)
    gd.alignment = CENTER
    gd.number_format = PCT0
    _box(ws, 5, 7, 6, 7, OV_INPUT_BORDER, "medium")

    # KPI cards (OCI totals are after the discount in $G$5).
    ws.merge_cells("B4:C4")
    band("B4", "Total Monthly Cost")
    ws.merge_cells("B5:C6")
    # Added services roll into the monthly + annual totals. Native OCI services take the
    # OCI discount ($G$5); 3rd-party licensing (Windows / SQL Server) is billed at list.
    oci = round(float(extra_oci or 0), 2)
    tp = round(float(extra_third_party or 0), 2)
    m_extra = (f"+{oci}*(1-$G$5)" if oci else "") + (f"+{tp}" if tp else "")
    a_extra = (f"+{round(oci * 12, 2)}*(1-$G$5)" if oci else "") + (f"+{round(tp * 12, 2)}" if tp else "")
    ws["B5"] = f"={BOM}!J15*(1-$G$5){m_extra}"
    ws.merge_cells("E4:F4")
    band("E4", "Total Annual Cost (Full)")
    ws.merge_cells("E5:F6")
    ws["E5"] = f"={BOM}!J18*(1-$G$5){a_extra}"
    for cell in ("B4", "E4"):
        ws[cell].fill = _fill(OV_BLUE)
    for cell in ("B5", "E5"):
        c = ws[cell]
        c.fill = _fill(OV_LIGHT)
        c.font = Font(bold=True, size=24)
        c.alignment = CENTER
        c.number_format = MONEY2
    _box(ws, 5, 2, 6, 3, OV_NAVY, "medium")
    _box(ws, 5, 5, 6, 6, OV_NAVY, "medium")

    # 5-Year Cost Projection (Utilization % comes from the app ramp)
    ws.merge_cells("A8:D8")
    band("A8", "5-Year Cost Projection")
    for cell, text in (("A9", "Year"), ("B9", "Utilization %"), ("C9", "OCI Annual Cost"), ("D9", "Cumulative")):
        colhead(cell, text)
    for y in range(5):
        row = 10 + y
        ws[f"A{row}"] = y + 1
        ws[f"A{row}"].alignment = CENTER
        util = util_by_year[y] if y < len(util_by_year) else 1.0
        b = ws[f"B{row}"]
        b.value = round(util, 4)
        b.number_format = PCT0
        b.alignment = CENTER
        b.fill = _fill(OV_INPUT)
        ws[f"C{row}"] = f"=$E$5*B{row}"
        ws[f"D{row}"] = (f"=C{row}" if y == 0 else f"=D{row - 1}+C{row}")
        ws[f"C{row}"].number_format = MONEY2
        ws[f"D{row}"].number_format = MONEY2
    ws.merge_cells("A15:B15")
    ws["A15"] = "5-Year Total"
    ws["C15"] = "=SUM(C10:C14)"
    ws["D15"] = "=D14"
    for cell in ("A15", "B15", "C15", "D15"):
        ws[cell].font = Font(bold=True)
        ws[cell].fill = _fill(OV_LIGHT)
    ws["C15"].number_format = MONEY2
    ws["D15"].number_format = MONEY2
    _box(ws, 10, 2, 14, 2, OV_INPUT_BORDER, "thin")
    _box(ws, 9, 1, 15, 4, OV_NAVY, "thin")

    # Existing vs OCI savings
    ws.merge_cells("A17:D17")
    band("A17", "Existing Infrastructure vs. OCI Estimate")
    rows = [
        ("A18", existing_label, "B18", existing_infra_cost, MONEY2),
        ("A19", "OCI Estimated Annual:", "B19", "=$E$5", MONEY2),
        ("A20", "Estimated Annual Savings:", "B20", "=B18-B19", MONEY2),
        ("A21", "Estimated 5-Year Savings:", "B21", "=SUM(U10:U14)-SUM(C10:C14)", MONEY2),
        ("A22", "Savings % vs. Existing:", "B22", "=IF(B18=0,0,(B18-B19)/B18)", "0.0%"),
    ]
    for la, label, ba, val, fmt in rows:
        ws[la] = label
        ws[la].font = Font(bold=True)
        c = ws[ba]
        c.value = val
        c.number_format = fmt
        c.alignment = CENTER
    ws["B18"].fill = _fill(OV_INPUT)
    _box(ws, 18, 2, 18, 2, OV_INPUT_BORDER, "medium")
    # Green (bold) font whenever savings are positive
    ws.conditional_formatting.add(
        "B20:B22",
        CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="FF008000", bold=True)),
    )

    # Chart data — comparison table
    ws.merge_cells("A23:F23")
    band("A23", "Chart Data — Comparison", size=12)
    for cell, text in (("A24", "Year"), ("B24", "OCI Estimate"), ("C24", "Existing Infra Cost"),
                       ("D24", "Combined Cost"), ("E24", "Current Spend"), ("F24", "Total Savings")):
        colhead(cell, text)
    ws["A25"] = "Today"
    ws["A25"].alignment = CENTER
    ws["C25"] = "=$B$18"
    ws["E25"] = "=$B$18"
    ws["F25"] = "=$B$18-C25"
    for y in range(5):
        r = 26 + y
        ws[f"A{r}"] = f"Year {y + 1}"
        ws[f"A{r}"].alignment = CENTER
        ws[f"B{r}"] = f"=C{10 + y}"
        ws[f"C{r}"] = f"=U{10 + y}"
        ws[f"D{r}"] = f"=B{r}+C{r}"
        # Current Spend tracks the AWS/Azure existing-spend ramp (cols T:V), not a flat value.
        ws[f"E{r}"] = f"=U{10 + y}"
        ws[f"F{r}"] = f"=E{r}-D{r}"
    for r in range(25, 31):
        for col in "BCDEF":
            ws[f"{col}{r}"].number_format = MONEY2
    _box(ws, 24, 1, 30, 6, OV_NAVY, "thin")
    # Total Savings: green when positive, red when negative.
    ws.conditional_formatting.add(
        "F25:F30",
        CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="FF008000", bold=True)),
    )
    ws.conditional_formatting.add(
        "F25:F30",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color="FFC00000", bold=True)),
    )

    # Existing-spend ramp adjuster (cols T:V)
    for col in "TUV":
        ws.column_dimensions[col].width = 18
    ws.merge_cells("T8:V8")
    band("T8", "Existing Spend — Ramp %", size=12)
    for cell, text in (("T9", "Year"), ("U9", "Existing Infra Cost"), ("V9", "Existing Util %")):
        colhead(cell, text)
    for y in range(5):
        r = 10 + y
        ws[f"T{r}"] = f"Year {y + 1}"
        ws[f"T{r}"].alignment = CENTER
        v = ws[f"V{r}"]
        v.value = 1
        v.number_format = PCT0
        v.alignment = CENTER
        v.fill = _fill(OV_INPUT)
        ws[f"U{r}"] = f"=$B$18*V{r}"
        ws[f"U{r}"].number_format = MONEY2
    _box(ws, 10, 22, 14, 22, OV_INPUT_BORDER, "thin")
    _box(ws, 9, 20, 14, 22, OV_NAVY, "thin")

    # Comparison chart (clustered columns + reference line)
    # Stacked columns: OCI Estimate (red) + Existing Infra Cost (grey) stacked, with
    # the Current Spend reference line across the top.
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.varyColors = False
    chart.title = "Existing Infra vs. OCI Estimate (Today + 5-Year)"
    chart.height = 9
    chart.width = 20
    chart.legend.position = "b"  # legend along the bottom, like the reference
    chart.legend.overlay = False
    # B = OCI Estimate, C = Existing Infra Cost (skip D = Combined; the stack is the total).
    data = Reference(ws, min_col=2, max_col=3, min_row=24, max_row=30)
    cats = Reference(ws, min_col=1, min_row=25, max_row=30)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # Set colors on the series AND every data point so Excel can't fall back to a theme.
    series_colors = ["C00000", "808080"]  # OCI = red, Existing = grey
    for idx, color in enumerate(series_colors):
        if idx >= len(chart.series):
            continue
        ser = chart.series[idx]
        ser.graphicalProperties = GraphicalProperties(solidFill=color)
        ser.dPt = []
        for pt in range(6):  # Today + Year 1..5
            dp = DataPoint(idx=pt)
            dp.graphicalProperties = GraphicalProperties(solidFill=color)
            ser.dPt.append(dp)

    line = LineChart()
    line.varyColors = False
    ldata = Reference(ws, min_col=5, max_col=5, min_row=24, max_row=30)
    line.add_data(ldata, titles_from_data=True)
    # Current Spend = faded red reference line
    if line.series:
        lgp = GraphicalProperties()
        lgp.line.solidFill = "E6A9A9"
        lgp.line.width = 28000  # ~2.2pt so the reference line reads clearly
        line.series[0].graphicalProperties = lgp
        line.series[0].smooth = False
    chart += line
    ws.add_chart(chart, "H8")

    # Small disclaimer note.
    dcell = ws["A32"]
    dcell.value = ("*Budgetary estimate only — not a quote. The OCI discount shown is "
                   "applied to the OCI totals on this Overview; detailed sheets list OCI list pricing.")
    dcell.font = Font(italic=True, size=9, color="FF808080")


def build_cloud_overview_sheet(ws, util_by_year, oci_monthly, existing_monthly,
                               existing_label="Current AWS Spend (monthly):", oci_discount=0.0):
    """Render the 'Overview' sheet for the AWS->OCI bill comparison.

    Mirrors the on-prem `build_overview_sheet` layout, red theme, and stacked
    column + Current Spend line chart exactly, but feeds the figures straight
    from the cloud pricing payload instead of cross-sheet BOM references:
      - oci_monthly     = OCI monthly total (discounted if a discount is set),
                          ties to the 'Product Breakdown ' OCI total.
      - existing_monthly = current AWS spend (sourceMonthlyCost), the grey
                          'Existing'/'Current Spend' series.
    util_by_year is the same 5-year ramp the on-prem Overview consumes.
    """
    oci_monthly = float(oci_monthly or 0)
    existing_monthly = float(existing_monthly or 0)
    oci_annual = oci_monthly * 12.0
    existing_annual = existing_monthly * 12.0

    ws.column_dimensions["A"].width = 28
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 16

    def band(cell, text, size=13):
        c = ws[cell]
        c.value = text
        c.fill = _fill(OV_NAVY)
        c.font = Font(bold=True, color="FFFFFFFF", size=size)
        c.alignment = CENTER

    def colhead(cell, text):
        c = ws[cell]
        c.value = text
        c.fill = _fill(OV_BLUE)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.alignment = CENTER

    # Title banner
    ws.merge_cells("A1:H2")
    t = ws["A1"]
    t.value = "AWS Bill to OCI Cost Overview"
    t.fill = _fill(OV_NAVY)
    t.font = Font(bold=True, color="FFFFFFFF", size=26, name="Calibri")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26

    # OCI discount applied from the app (the OCI figures are already net of it).
    band("G4", "OCI Discount*")
    ws["G4"].fill = _fill(OV_BLUE)
    gd = ws["G5"]
    gd.value = round(float(oci_discount or 0), 4)
    gd.fill = _fill(OV_INPUT)
    gd.font = Font(bold=True, size=18)
    gd.alignment = CENTER
    gd.number_format = PCT0
    _box(ws, 5, 7, 6, 7, OV_INPUT_BORDER, "medium")

    # KPI cards (OCI total is already discounted)
    ws.merge_cells("B4:C4")
    band("B4", "OCI Total Monthly Cost")
    ws.merge_cells("B5:C6")
    ws["B5"] = oci_monthly
    ws.merge_cells("E4:F4")
    band("E4", "OCI Total Annual Cost")
    ws.merge_cells("E5:F6")
    ws["E5"] = "=B5*12"
    for cell in ("B4", "E4"):
        ws[cell].fill = _fill(OV_BLUE)
    for cell in ("B5", "E5"):
        c = ws[cell]
        c.fill = _fill(OV_LIGHT)
        c.font = Font(bold=True, size=24)
        c.alignment = CENTER
        c.number_format = MONEY2
    _box(ws, 5, 2, 6, 3, OV_NAVY, "medium")
    _box(ws, 5, 5, 6, 6, OV_NAVY, "medium")

    # 5-Year Cost Projection (Utilization % comes from the app ramp)
    ws.merge_cells("A8:D8")
    band("A8", "5-Year Cost Projection")
    for cell, text in (("A9", "Year"), ("B9", "Utilization %"), ("C9", "OCI Annual Cost"), ("D9", "Cumulative")):
        colhead(cell, text)
    for y in range(5):
        row = 10 + y
        ws[f"A{row}"] = y + 1
        ws[f"A{row}"].alignment = CENTER
        util = util_by_year[y] if y < len(util_by_year) else 1.0
        b = ws[f"B{row}"]
        b.value = round(util, 4)
        b.number_format = PCT0
        b.alignment = CENTER
        b.fill = _fill(OV_INPUT)
        ws[f"C{row}"] = f"=$E$5*B{row}"
        ws[f"D{row}"] = (f"=C{row}" if y == 0 else f"=D{row - 1}+C{row}")
        ws[f"C{row}"].number_format = MONEY2
        ws[f"D{row}"].number_format = MONEY2
    ws.merge_cells("A15:B15")
    ws["A15"] = "5-Year Total"
    ws["C15"] = "=SUM(C10:C14)"
    ws["D15"] = "=D14"
    for cell in ("A15", "B15", "C15", "D15"):
        ws[cell].font = Font(bold=True)
        ws[cell].fill = _fill(OV_LIGHT)
    ws["C15"].number_format = MONEY2
    ws["D15"].number_format = MONEY2
    _box(ws, 10, 2, 14, 2, OV_INPUT_BORDER, "thin")
    _box(ws, 9, 1, 15, 4, OV_NAVY, "thin")

    # Existing (AWS) vs OCI savings
    ws.merge_cells("A17:D17")
    band("A17", "Current AWS Spend vs. OCI Estimate")
    rows = [
        ("A18", existing_label, "B18", existing_monthly, MONEY2),
        ("A19", "OCI Estimated Monthly:", "B19", "=$B$5", MONEY2),
        ("A20", "Estimated Monthly Savings:", "B20", "=B18-B19", MONEY2),
        ("A21", "Estimated 5-Year Savings:", "B21", "=SUM(U10:U14)-SUM(C10:C14)", MONEY2),
        ("A22", "Savings % vs. AWS:", "B22", "=IF(B18=0,0,(B18-B19)/B18)", "0.0%"),
    ]
    for la, label, ba, val, fmt in rows:
        ws[la] = label
        ws[la].font = Font(bold=True)
        c = ws[ba]
        c.value = val
        c.number_format = fmt
        c.alignment = CENTER
    ws["B18"].fill = _fill(OV_INPUT)
    _box(ws, 18, 2, 18, 2, OV_INPUT_BORDER, "medium")
    # Green (bold) font whenever savings are positive
    ws.conditional_formatting.add(
        "B20:B22",
        CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="FF008000", bold=True)),
    )

    # Chart data — comparison table (annual figures, ramped over 5 years)
    ws.merge_cells("A23:F23")
    band("A23", "Chart Data — Comparison", size=12)
    for cell, text in (("A24", "Year"), ("B24", "OCI Estimate"), ("C24", "Existing (AWS) Cost"),
                       ("D24", "Combined Cost"), ("E24", "Current Spend"), ("F24", "Total Savings")):
        colhead(cell, text)
    # Annualize the monthly AWS spend so the comparison table is apples-to-apples
    # with the OCI annual figures in cols C10:C14.
    ws["A25"] = "Today"
    ws["A25"].alignment = CENTER
    ws["C25"] = "=$B$18*12"
    ws["E25"] = "=$B$18*12"
    ws["F25"] = "=E25-C25"
    for y in range(5):
        r = 26 + y
        ws[f"A{r}"] = f"Year {y + 1}"
        ws[f"A{r}"].alignment = CENTER
        ws[f"B{r}"] = f"=C{10 + y}"
        ws[f"C{r}"] = f"=U{10 + y}"
        ws[f"D{r}"] = f"=B{r}+C{r}"
        # Current Spend tracks the existing-AWS-spend ramp (cols T:V), not a flat value.
        ws[f"E{r}"] = f"=U{10 + y}"
        ws[f"F{r}"] = f"=E{r}-D{r}"
    for r in range(25, 31):
        for col in "BCDEF":
            ws[f"{col}{r}"].number_format = MONEY2
    _box(ws, 24, 1, 30, 6, OV_NAVY, "thin")
    # Total Savings: green when positive, red when negative.
    ws.conditional_formatting.add(
        "F25:F30",
        CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="FF008000", bold=True)),
    )
    ws.conditional_formatting.add(
        "F25:F30",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color="FFC00000", bold=True)),
    )

    # Existing-spend ramp adjuster (cols T:V) — annual existing AWS cost per year.
    for col in "TUV":
        ws.column_dimensions[col].width = 18
    ws.merge_cells("T8:V8")
    band("T8", "Existing Spend — Ramp %", size=12)
    for cell, text in (("T9", "Year"), ("U9", "Existing (AWS) Cost"), ("V9", "Existing Util %")):
        colhead(cell, text)
    for y in range(5):
        r = 10 + y
        ws[f"T{r}"] = f"Year {y + 1}"
        ws[f"T{r}"].alignment = CENTER
        v = ws[f"V{r}"]
        v.value = 1
        v.number_format = PCT0
        v.alignment = CENTER
        v.fill = _fill(OV_INPUT)
        ws[f"U{r}"] = f"=$B$18*12*V{r}"
        ws[f"U{r}"].number_format = MONEY2
    _box(ws, 10, 22, 14, 22, OV_INPUT_BORDER, "thin")
    _box(ws, 9, 20, 14, 22, OV_NAVY, "thin")

    # Comparison chart (stacked columns + reference line) — identical to on-prem.
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.varyColors = False
    chart.title = "Current AWS Spend vs. OCI Estimate (Today + 5-Year)"
    chart.height = 9
    chart.width = 20
    chart.legend.position = "b"  # legend along the bottom, like the reference
    chart.legend.overlay = False
    # B = OCI Estimate, C = Existing (AWS) Cost (skip D = Combined; the stack is the total).
    data = Reference(ws, min_col=2, max_col=3, min_row=24, max_row=30)
    cats = Reference(ws, min_col=1, min_row=25, max_row=30)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # Set colors on the series AND every data point so Excel can't fall back to a theme.
    series_colors = ["C00000", "808080"]  # OCI = red, Existing = grey
    for idx, color in enumerate(series_colors):
        if idx >= len(chart.series):
            continue
        ser = chart.series[idx]
        ser.graphicalProperties = GraphicalProperties(solidFill=color)
        ser.dPt = []
        for pt in range(6):  # Today + Year 1..5
            dp = DataPoint(idx=pt)
            dp.graphicalProperties = GraphicalProperties(solidFill=color)
            ser.dPt.append(dp)

    line = LineChart()
    line.varyColors = False
    ldata = Reference(ws, min_col=5, max_col=5, min_row=24, max_row=30)
    line.add_data(ldata, titles_from_data=True)
    # Current Spend = faded red reference line
    if line.series:
        lgp = GraphicalProperties()
        lgp.line.solidFill = "E6A9A9"
        lgp.line.width = 28000  # ~2.2pt so the reference line reads clearly
        line.series[0].graphicalProperties = lgp
        line.series[0].smooth = False
    chart += line
    ws.add_chart(chart, "H8")

    # Small disclaimer note.
    dcell = ws["A32"]
    dcell.value = ("*Budgetary estimate only — not a quote. The OCI discount shown is "
                   "applied to the OCI totals on this Overview; detailed sheets list OCI list pricing.")
    dcell.font = Font(italic=True, size=9, color="FF808080")


def add_comparison_to_pricing_overview(ws, start_row, oci_monthly_ref, oci_annual_ref,
                                       aws_monthly, util_by_year,
                                       aws_ramp=(0.5, 0.0, 0.0, 0.0, 0.0)):
    """Render the Cloud Bill Overview's below-row-7 blocks (5-Year Cost Projection,
    AWS-vs-OCI savings, Existing-spend ramp %, Chart Data + chart) onto the Pricing
    Overview at `start_row`, wired to that sheet's live OCI total cells:
      oci_monthly_ref / oci_annual_ref  e.g. "$B$22" / "$B$23"  (same-sheet refs)
      aws_monthly  = current AWS spend, written as an editable input value
      util_by_year = OCI 5-year utilization ramp (list of 5 fractions)
      aws_ramp     = existing-AWS spend ramp per year (default 50% yr1, 0% after),
                     kept as live editable % cells feeding the Combined / Savings math.
    """
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.marker import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.formatting.rule import CellIsRule

    s = start_row

    def band(cell, text, size=13):
        c = ws[cell]
        c.value = text
        c.fill = _fill(OV_NAVY)
        c.font = Font(bold=True, color="FFFFFFFF", size=size)
        c.alignment = CENTER

    def colhead(cell, text):
        c = ws[cell]
        c.value = text
        c.fill = _fill(OV_BLUE)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.alignment = CENTER

    # ---- 5-Year Cost Projection (OCI utilization ramp) ----
    ws.merge_cells(f"A{s}:D{s}")
    band(f"A{s}", "5-Year Cost Projection")
    for col, text in (("A", "Year"), ("B", "Utilization %"), ("C", "OCI Annual Cost"), ("D", "Cumulative")):
        colhead(f"{col}{s+1}", text)
    for y in range(5):
        r = s + 2 + y
        ws[f"A{r}"] = y + 1
        ws[f"A{r}"].alignment = CENTER
        util = util_by_year[y] if y < len(util_by_year) else 1.0
        b = ws[f"B{r}"]
        b.value = round(util, 4); b.number_format = PCT0; b.alignment = CENTER; b.fill = _fill(OV_INPUT)
        ws[f"C{r}"] = f"={oci_annual_ref}*B{r}"
        ws[f"D{r}"] = (f"=C{r}" if y == 0 else f"=D{r-1}+C{r}")
        ws[f"C{r}"].number_format = MONEY2
        ws[f"D{r}"].number_format = MONEY2
    tr = s + 7
    ws.merge_cells(f"A{tr}:B{tr}")
    ws[f"A{tr}"] = "5-Year Total"
    ws[f"C{tr}"] = f"=SUM(C{s+2}:C{s+6})"
    ws[f"D{tr}"] = f"=D{s+6}"
    for col in "ABCD":
        ws[f"{col}{tr}"].font = Font(bold=True)
        ws[f"{col}{tr}"].fill = _fill(OV_LIGHT)
    ws[f"C{tr}"].number_format = MONEY2
    ws[f"D{tr}"].number_format = MONEY2
    _box(ws, s+2, 2, s+6, 2, OV_INPUT_BORDER, "thin")
    _box(ws, s+1, 1, tr, 4, OV_NAVY, "thin")

    # ---- Existing-AWS spend ramp % (editable; default 50% yr1, 0% after) ----
    ws.merge_cells(f"H{s}:J{s}")
    band(f"H{s}", "Existing Spend — Ramp %", size=12)
    for col, text in (("H", "Year"), ("I", "Existing (AWS) Cost"), ("J", "Existing Util %")):
        colhead(f"{col}{s+1}", text)
    aws_cell = f"$B${s+10}"  # the "Current AWS Spend (monthly)" input, defined below (row sv+1 = s+10)
    for y in range(5):
        r = s + 2 + y
        ws[f"H{r}"] = f"Year {y+1}"; ws[f"H{r}"].alignment = CENTER
        v = ws[f"J{r}"]
        rv = aws_ramp[y] if y < len(aws_ramp) else 0.0
        v.value = round(rv, 4); v.number_format = PCT0; v.alignment = CENTER; v.fill = _fill(OV_INPUT)
        ws[f"I{r}"] = f"={aws_cell}*12*J{r}"
        ws[f"I{r}"].number_format = MONEY2
    _box(ws, s+2, 10, s+6, 10, OV_INPUT_BORDER, "thin")
    _box(ws, s+1, 8, s+6, 10, OV_NAVY, "thin")

    # ---- Current AWS Spend vs. OCI Estimate ----
    sv = s + 9
    ws.merge_cells(f"A{sv}:D{sv}")
    band(f"A{sv}", "Current AWS Spend vs. OCI Estimate")
    rows_spec = [
        (f"A{sv+1}", "Current AWS Spend (monthly):", f"B{sv+1}", float(aws_monthly or 0), MONEY2),
        (f"A{sv+2}", "OCI Estimated Monthly:", f"B{sv+2}", f"={oci_monthly_ref}", MONEY2),
        (f"A{sv+3}", "Estimated Monthly Savings:", f"B{sv+3}", f"=B{sv+1}-B{sv+2}", MONEY2),
        (f"A{sv+4}", "Estimated 5-Year Savings:", f"B{sv+4}", f"=F{s+23}", MONEY2),
        (f"A{sv+5}", "Savings % vs. AWS:", f"B{sv+5}", f"=IF(B{sv+1}=0,0,(B{sv+1}-B{sv+2})/B{sv+1})", "0.0%"),
    ]
    for la, label, ba, val, fmt in rows_spec:
        ws[la] = label; ws[la].font = Font(bold=True)
        c = ws[ba]; c.value = val; c.number_format = fmt; c.alignment = CENTER
    ws[f"B{sv+1}"].fill = _fill(OV_INPUT)
    _box(ws, sv+1, 2, sv+1, 2, OV_INPUT_BORDER, "medium")
    ws.conditional_formatting.add(
        f"B{sv+3}:B{sv+5}",
        CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="FF008000", bold=True)),
    )

    # ---- Chart Data — Comparison ----
    cd = s + 16
    ws.merge_cells(f"A{cd}:F{cd}")
    band(f"A{cd}", "Chart Data — Comparison", size=12)
    for col, text in (("A", "Year"), ("B", "OCI Estimate"), ("C", "Existing (AWS) Cost"),
                      ("D", "Combined Cost"), ("E", "Current Spend"), ("F", "Total Savings")):
        colhead(f"{col}{cd+1}", text)
    ws[f"A{cd+2}"] = "Today"; ws[f"A{cd+2}"].alignment = CENTER
    ws[f"C{cd+2}"] = f"={aws_cell}*12"
    ws[f"E{cd+2}"] = f"={aws_cell}*12"
    ws[f"F{cd+2}"] = f"=E{cd+2}-C{cd+2}"
    for y in range(5):
        r = cd + 3 + y
        ws[f"A{r}"] = f"Year {y+1}"; ws[f"A{r}"].alignment = CENTER
        ws[f"B{r}"] = f"=C{s+2+y}"          # OCI Estimate (annual, ramped by utilization)
        ws[f"C{r}"] = f"=I{s+2+y}"          # Existing (AWS) still running — reflects the ramp
        ws[f"D{r}"] = f"=B{r}+C{r}"         # Combined migration cost = OCI + remaining AWS
        ws[f"E{r}"] = f"={aws_cell}*12"     # Current Spend = STATIC full annual AWS (do-nothing baseline)
        # Total Savings is CUMULATIVE — each year adds that year's (Current Spend - Combined).
        ws[f"F{r}"] = f"=F{r-1}+(E{r}-D{r})"
    for r in range(cd+2, cd+8):
        for col in "BCDEF":
            ws[f"{col}{r}"].number_format = MONEY2
    _box(ws, cd+1, 1, cd+7, 6, OV_NAVY, "thin")
    ws.conditional_formatting.add(
        f"F{cd+2}:F{cd+7}",
        CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="FF008000", bold=True)))
    ws.conditional_formatting.add(
        f"F{cd+2}:F{cd+7}",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color="FFC00000", bold=True)))

    # ---- Comparison chart (stacked columns + Current Spend reference line) ----
    chart = BarChart()
    chart.type = "col"; chart.grouping = "stacked"; chart.overlap = 100; chart.varyColors = False
    chart.title = "Current AWS Spend vs. OCI Estimate (Today + 5-Year)"
    chart.height = 9; chart.width = 20
    chart.legend.position = "b"; chart.legend.overlay = False
    data = Reference(ws, min_col=2, max_col=3, min_row=cd+1, max_row=cd+7)
    cats = Reference(ws, min_col=1, min_row=cd+2, max_row=cd+7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for idx, color in enumerate(["C00000", "808080"]):
        if idx >= len(chart.series):
            continue
        ser = chart.series[idx]
        ser.graphicalProperties = GraphicalProperties(solidFill=color)
        ser.dPt = []
        for pt in range(6):
            dp = DataPoint(idx=pt)
            dp.graphicalProperties = GraphicalProperties(solidFill=color)
            ser.dPt.append(dp)
    line = LineChart(); line.varyColors = False
    ldata = Reference(ws, min_col=5, max_col=5, min_row=cd+1, max_row=cd+7)
    line.add_data(ldata, titles_from_data=True)
    if line.series:
        lgp = GraphicalProperties(); lgp.line.solidFill = "E6A9A9"; lgp.line.width = 28000
        line.series[0].graphicalProperties = lgp; line.series[0].smooth = False
    chart += line
    ws.add_chart(chart, f"H{sv}")


_CLOUD_GROUP_MAP = {
    "compute": "Compute", "containers": "Compute",
    "containers and functions": "Compute",
    "storage": "Storage",
    "database": "Database", "oracle databases": "Database",
    "databases": "Database", "analytics": "Database",
    "networking": "Networking",
    "security and identity": "Security", "security": "Security",
    "identity": "Security", "security and compliance": "Security",
    "observability and management": "Obs. & Management",
    "obs. & management": "Obs. & Management",
    "management": "Obs. & Management", "monitoring": "Obs. & Management",
    "application integration": "DevOps",
    "ai and machine learning": "AI & Machine Learning",
    "ai & machine learning": "AI & Machine Learning",
    "artificial intelligence": "AI & Machine Learning",
    "devops": "DevOps",
    "developer tools": "DevOps", "developer services": "DevOps",
    "end user computing": "Compute",
    "support": "Support",
    "marketplace": "Marketplace",
    "other services": "Other Services", "other": "Other Services",
}


# Reference AWS->OCI->product-group lookup (extracted from the reference's
# "Service Comp List" sheet). This is the authoritative product grouping the
# reference uses, so the export groups identically (e.g. Savings Plans -> Compute).
def _load_service_comp_list():
    path = os.path.join(os.path.dirname(__file__), "data", "service_comp_list.json")
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except (OSError, ValueError):
        return {}, []
    by_norm = {}
    for e in data.get("entries", []):
        key = e.get("awsNorm")
        if key and key not in by_norm:
            by_norm[key] = e
    # Sort norm keys longest-first for greedy substring matching.
    keys_by_len = sorted(by_norm.keys(), key=len, reverse=True)
    return by_norm, keys_by_len


_SERVICE_COMP_BY_NORM, _SERVICE_COMP_KEYS = _load_service_comp_list()


def _norm_service(s):
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def _ref_lookup(aws_service):
    """Return the reference entry {group, free, ociEquivalent} for an AWS service
    name, matching the reference's VLOOKUP (exact first, then best substring)."""
    norm = _norm_service(aws_service)
    if not norm:
        return None
    hit = _SERVICE_COMP_BY_NORM.get(norm)
    if hit:
        return hit
    # Best substring match: a known service name contained in the bill's label.
    for key in _SERVICE_COMP_KEYS:
        if len(key) >= 4 and key in norm:
            return _SERVICE_COMP_BY_NORM[key]
    return None


def _cloud_product_group(category, aws_service, oci_product=None):
    cat_l = (category or "").strip().lower()
    prod_l = (oci_product or "").lower()
    # GPU / AI compute always rolls up under AI & Machine Learning, even though the
    # underlying AWS service (EC2) would otherwise ref-lookup to Compute.
    if ("gpu" in prod_l
            or cat_l in ("ai & machine learning", "ai and machine learning",
                         "artificial intelligence")):
        return "AI & Machine Learning"
    ref = _ref_lookup(aws_service)
    if ref and ref.get("group"):
        return ref["group"]
    svc = (aws_service or "").lower()
    if "support" in svc:
        return "Support"
    if "marketplace" in svc:
        return "Marketplace"
    return _CLOUD_GROUP_MAP.get((category or "").strip().lower(), "Other Services")


# ============================================================================
# Cloud-bill AWS -> OCI comparison workbook (reference: "Product Breakdown ")
# ============================================================================

# Group ordering used everywhere (matches the reference's 11 product groups).
_CLOUD_GROUP_ORDER = [
    "Compute", "Database", "Storage", "Networking", "Support",
    "Obs. & Management", "Other Services", "Security",
    "AI & Machine Learning", "DevOps", "Marketplace",
]

# Per-group fill colors for the conditional formatting on the Product Group cells.
_CLOUD_GROUP_COLORS = {
    "Compute": "D9E1F2",
    "Database": "FF9F9F",
    "Storage": "E2EFDA",
    "Networking": "DDEBF7",
    "Security": "A6C56B",
    "Obs. & Management": "9FE1E1",
    "AI & Machine Learning": "FF99CC",
    "DevOps": "E68C42",
    "Other Services": "FFF2CC",
    "Support": "D9D9D9",
    "Marketplace": "B493E1",
}

# Number formats (match the reference exactly).
_ACCT2 = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
_ACCT0 = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
_PCT = "0%"
_MONEY0 = '"$"#,##0'

# Reference colors (ARGB without the leading "FF").
_C_BLUE = "5B9BD5"       # AWS banner / "AWS Cost" header
_C_GOLD = "FFC000"       # USD banners
_C_RED = "FF5050"        # OCI banner
_C_RED2 = "FF6165"       # OCI Cost legend header
_C_HDR = "0070C0"        # data header row
_C_DATA = "DDEBF7"       # data-cell light blue
_C_TOTAL = "8EA9DB"      # totals row mid blue
_C_GRAY = "D9D9D9"       # discount tier / banner gray
_C_GREEN_HDR = "92D050"  # legend green headers/totals
_C_BOX_GREEN = "E2EFDA"  # savings box 1 pale green
_C_BOX_GOLD = "FFF2CC"   # savings box 2 pale gold
_C_CARRY = "4472C4"      # carried-cost blue font
_C_GOOD_FONT = "006100"  # "Good" green font
_C_GOOD_FILL = "C6EFCE"  # "Good" green fill

# AWS services we treat as free in OCI when their computed OCI cost is 0.
_CLOUD_FREE_SERVICES = {
    "amazon virtual private cloud", "aws key management service",
    "aws identity and access management", "amazon cloudwatch",
    "aws cloudtrail", "aws config", "amazon sns",
    "amazon simple notification service",
}

_THIN = Side(style="thin", color="FF000000")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_VCENTER = Alignment(vertical="center")
_CTR = Alignment(horizontal="center", vertical="center")
_CTR_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)


WORKFLOW_SHEET = "_workflow"
_WORKFLOW_CHUNK = 30000


def embed_workflow_state(wb, workflow_json):
    """Embed the full app workflow state (a JSON string) in a hidden sheet so the
    exported workbook can be re-imported to recreate the app window exactly.
    The JSON is chunked across column A (cell text limit is 32767 chars)."""
    if not workflow_json:
        return
    ws = wb.create_sheet(WORKFLOW_SHEET)
    ws.sheet_state = "hidden"
    ws["A1"] = "OCI BOM workflow state — do not edit. Re-import this file to restore your session."
    text = workflow_json if isinstance(workflow_json, str) else json.dumps(workflow_json)
    row = 2
    for i in range(0, len(text), _WORKFLOW_CHUNK):
        ws.cell(row=row, column=1, value=text[i:i + _WORKFLOW_CHUNK])
        row += 1


def read_workflow_state(path):
    """Extract embedded workflow JSON from an exported .xlsx (the hidden
    _workflow sheet) or from a raw .json file. Returns a dict or None."""
    p = str(path).lower()
    if p.endswith(".json"):
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    if WORKFLOW_SHEET not in wb.sheetnames:
        return None
    ws = wb[WORKFLOW_SHEET]
    parts = []
    for r, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
        if r == 1:
            continue  # header note
        if row and row[0]:
            parts.append(str(row[0]))
    text = "".join(parts)
    return json.loads(text) if text else None


def add_cloud_comparison_sheets(wb, pricing, ramp=None, bom_name="", oci_discount=0.0,
                                extra_services=None, hours=HOURS, use_active=False):
    """Add the AWS->OCI bill-comparison sheets (Product Breakdown, Service Mapping,
    Notes + Assumptions, Overview) to an existing workbook.

    `use_active=True` renders Product Breakdown onto wb.active (for a standalone workbook);
    otherwise all four are new sheets (for appending onto the Full BOM template so no bill
    data or mappings are lost). Returns the Product Breakdown worksheet."""
    try:
        oci_discount = float(oci_discount or 0)
    except (TypeError, ValueError):
        oci_discount = 0.0
    oci_discount = min(max(oci_discount, 0.0), 1.0)
    rows = pricing.get("rows", []) or []

    agg = {}
    for r in rows:
        if (r.get("costAction") or "") == "remove":
            continue
        aws_svc = r.get("sourceService") or "Other"
        oci_prod = r.get("ociProduct") or "Needs review"
        group = _cloud_product_group(r.get("ociServiceCategory"), aws_svc, oci_prod)
        key = (group, aws_svc, oci_prod)
        a = agg.setdefault(key, {"aws": 0.0, "oci": 0.0, "carry": False})
        a["aws"] += float(r.get("sourceMonthlyCost") or 0)
        a["oci"] += float(r.get("monthly") or 0)
        if (r.get("costAction") or "") == "carry":
            a["carry"] = True

    by_group = {}
    for (group, aws_svc, oci_prod), v in agg.items():
        by_group.setdefault(group, []).append((aws_svc, oci_prod, v))
    ordered = []
    present_groups = []
    for group in _CLOUD_GROUP_ORDER:
        if group not in by_group:
            continue
        present_groups.append(group)
        for aws_svc, oci_prod, v in sorted(by_group[group], key=lambda x: -x[2]["aws"]):
            ordered.append((group, aws_svc, oci_prod, v))

    totals = pricing.get("totals", {}) or {}
    oci_monthly = 0.0
    for r in rows:
        if (r.get("costAction") or "") == "remove":
            continue
        m = float(r.get("monthly") or 0)
        if (r.get("costAction") or "") == "carry" or not r.get("monthly"):
            oci_monthly += m
        else:
            oci_monthly += m * (1.0 - oci_discount)
    if not rows:
        oci_monthly = float(totals.get("monthly") or 0)

    if extra_services:
        import oci_catalog
        priced, _ = oci_catalog.price_extras(extra_services, hours)
        for s in priced:
            grp = "Added OCI Services"
            if grp not in present_groups:
                present_groups.append(grp)
            third = bool(s.get("thirdParty"))
            ordered.append((grp, s.get("sizing") or "Added in app", s["name"],
                            {"aws": 0.0, "oci": float(s["monthly"] or 0), "carry": third}))
            m = float(s["monthly"] or 0)
            oci_monthly += m if third else m * (1.0 - oci_discount)

    # 3rd-party Windows OS licensing. On OCI, Windows is a separate SKU (AWS bakes it into
    # the instance rate), so it MUST be added to the OCI side for a fair comparison. It's
    # already 0 when the Hide Windows toggle is on, so summing it here honors that toggle.
    windows_total = sum(float(r.get("windowsLicenseMonthly") or 0)
                        for r in rows if (r.get("costAction") or "") != "remove")
    if windows_total > 0:
        grp = "3rd-Party Licensing"
        if grp not in present_groups:
            present_groups.append(grp)
        # carry=False so the OCI cost is written as a real value (a carried row would force
        # OCI = AWS cost, which is 0 for this line since AWS bundles Windows into compute).
        ordered.append((grp, "Windows OS licensing", "OCI Windows OS License (per OCPU-hr)",
                        {"aws": 0.0, "oci": windows_total, "carry": False}))
        oci_monthly += windows_total

    existing_monthly = float(totals.get("sourceMonthlyCost") or 0)

    pb = wb.active if use_active else wb.create_sheet("Product Breakdown ")
    _cloud_product_breakdown(pb, ordered, present_groups, bom_name, oci_discount)
    _cloud_service_mapping_sheet(wb.create_sheet("Service Mapping"), rows, oci_discount)
    _cloud_notes_sheet(wb.create_sheet("Notes + Assumptions"))
    build_cloud_overview_sheet(
        wb.create_sheet("Overview" if use_active else "Cloud Bill Overview"),
        _util_by_year(ramp), oci_monthly, existing_monthly, oci_discount=oci_discount,
    )
    return pb


def build_cloud_comparison_bytes(pricing, ramp=None, bom_name="", oci_discount=0.0, workflow_json=None, extra_services=None, hours=HOURS):
    """AWS -> OCI bill comparison workbook that reproduces the reference's
    'Product Breakdown ' printout. OCI prices are written as values; no
    cross-sheet pricing engine is needed.

    oci_discount is a fraction (0.0-1.0) set by the user near the ramp graph.
    The reference's discount logic is preserved (Total - Discounted column =
    list * (1 - discount)); at 0 the discounted column equals the app's OCI
    total exactly."""
    wb = Workbook()
    add_cloud_comparison_sheets(wb, pricing, ramp, bom_name, oci_discount,
                               extra_services, hours, use_active=True)
    embed_workflow_state(wb, workflow_json)
    wb.active = 0  # keep "Product Breakdown " as the active/first sheet
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cloud_product_breakdown(ws, ordered, present_groups, bom_name, oci_discount=0.0):
    ws.title = "Product Breakdown "  # trailing space is intentional

    # ---- sheet view / page setup ----
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 80
    ws.page_setup.orientation = "portrait"

    # ---- column widths ----
    widths = {"A": 4.57, "B": 27.29, "C": 50.71, "D": 18.71, "E": 22.14,
              "F": 21.71, "G": 17.57, "H": 9.43, "I": 62.71, "J": 16.71,
              "K": 25.43, "L": 25.43, "M": 16.29, "N": 33.43, "O": 12.0,
              "P": 17.43, "Q": 17.43, "R": 17.86, "S": 18.14, "T": 13.29,
              "U": 15.29}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    N = len(ordered)
    first = 4
    last = first + N - 1 if N else first - 1   # if N==0, last < first
    total_row = (last + 1) if N else first
    annual_row = total_row + 1
    discount_row = annual_row + 2
    box1_hdr = discount_row + 2
    box1_save = box1_hdr + 1
    box1_pct = box1_save + 1
    box2_hdr = box1_pct + 2
    box2_save = box2_hdr + 1
    box2_pct = box2_save + 1

    # The Cloud Service Product Group summary (Monthly + Annual) is placed BELOW the
    # main table, a few rows under the savings boxes.
    pg = len(present_groups)
    summary_start = box2_pct + 3
    summary_bottom = summary_start + 5 + 2 * (pg + 1) + 3

    # ---- row heights ----
    for r in range(1, 32):
        ws.row_dimensions[r].height = 18.0
    for r in range(32, max(51, summary_bottom + 2)):
        ws.row_dimensions[r].height = 18.75

    base_font = Font(name="Calibri", size=14)

    def cell(coord, value=None):
        c = ws[coord]
        if value is not None:
            c.value = value
        c.font = base_font
        return c

    def banner(coord, value, fill, *, font_color="000000", bold=True,
               wrap=False, border_all=False):
        c = ws[coord]
        c.value = value
        c.fill = _fill(fill)
        c.font = Font(name="Calibri", size=14, bold=bold, color="FF" + font_color)
        c.alignment = _CTR_WRAP if wrap else _CTR
        if border_all:
            c.border = _THIN_BORDER
        return c

    # ---- title / banner block (row 2) ----
    title = (bom_name.strip() if bom_name else "") or "AWS to OCI Bill Comparison"
    b2 = banner("B2", title, _C_BLUE, font_color="FFFFFF")
    b2.border = Border(left=_THIN, top=_THIN, bottom=_THIN)
    banner("C2", "AWS", _C_BLUE, font_color="FFFFFF")
    ws.merge_cells("D2:G2")
    banner("D2", "USD", _C_GOLD, border_all=True)
    ws["D2"].number_format = _ACCT2
    banner("I2", "OCI", _C_RED, font_color="FFFFFF")
    ws.merge_cells("J2:L2")
    banner("J2", "USD ", _C_GOLD, wrap=True)

    # ---- data header (row 3) ----
    hdrs = {"B3": "Product Group", "C3": "AWS Service", "D3": "List Costs",
            "E3": "Discounts", "F3": "Invoice Costs", "G3": "% of Total",
            "I3": "Offer Name", "J3": "Total - List",
            "K3": "Total - Discounted", "L3": "Total Savings"}
    for coord, text in hdrs.items():
        c = banner(coord, text, _C_HDR, font_color="FFFFFF", border_all=True)
        c.alignment = _CTR

    data_fill = _fill(_C_DATA)

    def data_cell(col, r, value=None, fmt=None, bold=False, font_color=None):
        c = ws[f"{col}{r}"]
        if value is not None:
            c.value = value
        c.font = Font(name="Calibri", size=14, bold=bold,
                      color=("FF" + font_color) if font_color else None)
        c.fill = data_fill
        c.border = _THIN_BORDER
        c.alignment = _VCENTER
        if fmt:
            c.number_format = fmt
        return c

    # ---- data rows ----
    for i, (group, aws_svc, oci_prod, v) in enumerate(ordered):
        r = first + i
        carried = bool(v["carry"])
        aws_cost = round(v["aws"], 2)
        oci_cost = round(v["oci"], 2)
        # The app already encodes free-on-OCI services as $0; mark any such line
        # FREE so the export's OCI total matches the app exactly.
        free = (not carried) and oci_cost == 0

        data_cell("B", r, group)                       # Product Group (value)
        data_cell("C", r, aws_svc)                     # AWS Service
        data_cell("D", r, f"=F{r}-E{r}", _ACCT2)       # List Costs = Invoice - Disc
        data_cell("E", r, None, _ACCT2)                # Discounts (blank)
        data_cell("F", r, aws_cost, _ACCT2)            # Invoice Costs (value)
        data_cell("G", r, f'=IFERROR(F{r}/$F${total_row},"")', _PCT)

        # Offer Name (bold). Carried/flagged -> "*" prefix + blue font.
        offer = ("*" + oci_prod) if carried else oci_prod
        data_cell("I", r, offer, bold=True,
                  font_color=(_C_CARRY if carried else None))

        # Total - List
        if free:
            data_cell("J", r, "FREE", _ACCT2)
        elif carried:
            data_cell("J", r, f"=D{r}", _ACCT2)
        else:
            data_cell("J", r, oci_cost, _ACCT2)

        # Total - Discounted
        if carried:
            data_cell("K", r, f"=F{r}", _ACCT2)
        else:
            data_cell("K", r,
                      f'=IF(J{r}="FREE","FREE",J{r}*(1-$J${discount_row}))',
                      _ACCT2)

        # Total Savings
        data_cell("L", r, f'=F{r}-IF(K{r}="FREE",0,K{r})', _ACCT2)

    # ---- totals row ----
    sum_lo, sum_hi = first, last
    totfill = _fill(_C_TOTAL)

    def tot_cell(col, r, value, fmt):
        c = ws[f"{col}{r}"]
        c.value = value
        c.fill = totfill
        c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
        c.alignment = _CTR
        c.border = _THIN_BORDER
        c.number_format = fmt
        return c

    rng = f"{sum_lo}:{sum_hi}" if N else f"{first}:{first}"
    tot_cell("C", total_row, "Total Costs - Monthly", _ACCT2)
    tot_cell("D", total_row, f"=SUM(D{sum_lo}:D{sum_hi})" if N else "=0", _ACCT0)
    tot_cell("E", total_row, f"=SUM(E{sum_lo}:E{sum_hi})" if N else "=0", _ACCT2)
    tot_cell("F", total_row, f"=SUM(F{sum_lo}:F{sum_hi})" if N else "=0", _ACCT0)
    tot_cell("G", total_row, f"=SUM(G{sum_lo}:G{sum_hi})" if N else "=0", _PCT)
    tot_cell("I", total_row, "Total Costs - Monthly", _ACCT2)
    tot_cell("J", total_row, f"=SUM(J{sum_lo}:J{sum_hi})" if N else "=0", _ACCT0)
    tot_cell("K", total_row, f"=SUM(K{sum_lo}:K{sum_hi})" if N else "=0", _ACCT0)
    tot_cell("L", total_row, f"=SUM(L{sum_lo}:L{sum_hi})" if N else "=0", _ACCT0)

    # ---- annual row ----
    tr = total_row
    tot_cell("C", annual_row, "Total Costs - Annual", _ACCT2)
    tot_cell("D", annual_row, f"=D{tr}*12", _ACCT0)
    tot_cell("E", annual_row, f"=E{tr}*12", _ACCT2)
    tot_cell("F", annual_row, f"=F{tr}*12", _ACCT0)
    tot_cell("I", annual_row, "Total Costs - Annual", _ACCT2)
    tot_cell("J", annual_row, f"=J{tr}*12", _ACCT0)
    tot_cell("K", annual_row, f"=K{tr}*12", _ACCT0)
    tot_cell("L", annual_row, f"=L{tr}*12", _ACCT0)

    # ---- discount tier ----
    gray = _fill(_C_GRAY)
    idr = cell(f"I{discount_row}", "OCI Discount")
    idr.fill = gray
    idr.font = Font(name="Calibri", size=14, bold=True)
    jdr = ws[f"J{discount_row}"]
    # User-set OCI discount (from the control near the ramp graph). The reference's
    # discount logic is kept intact (Total - Discounted = list * (1 - this cell));
    # at 0 the discounted column equals the app's OCI total exactly.
    jdr.value = round(float(oci_discount or 0), 4)
    jdr.fill = gray
    jdr.font = Font(name="Calibri", size=14, bold=True)
    jdr.number_format = _PCT

    # ---- savings box 1: AWS Invoice vs OCI Discount ----
    green = _fill(_C_BOX_GREEN)

    def box_label(coord, text, fill):
        c = cell(coord, text)
        c.fill = fill
        c.font = Font(name="Calibri", size=14, bold=True)
        return c

    cc = cell(f"I{box1_hdr}", "AWS Invoice Price vs OCI Discount Price ")
    cc.font = Font(name="Calibri", size=14, bold=True)
    bh = cell(f"J{box1_hdr}", "Monthly"); bh.fill = green; bh.font = Font(name="Calibri", size=14, bold=True); bh.alignment = _CTR
    bh = cell(f"K{box1_hdr}", "Annual"); bh.fill = green; bh.font = Font(name="Calibri", size=14, bold=True); bh.alignment = _CTR
    box_label(f"I{box1_save}", "Saving per month with OCI", green)
    c = cell(f"J{box1_save}", f"=F{tr}-K{tr}"); c.fill = green; c.number_format = _ACCT0
    c = cell(f"K{box1_save}", f"=J{box1_save}*12"); c.fill = green; c.number_format = _ACCT0
    box_label(f"I{box1_pct}", "Total % Savings", green)
    c = cell(f"J{box1_pct}", f"=IFERROR(J{box1_save}/F{tr},0)"); c.fill = green; c.number_format = _PCT

    # ---- savings box 2: AWS List vs OCI List ----
    gold = _fill(_C_BOX_GOLD)
    cc = cell(f"I{box2_hdr}", "AWS List Price vs OCI List Price ")
    cc.font = Font(name="Calibri", size=14, bold=True)
    bh = cell(f"J{box2_hdr}", "Monthly"); bh.fill = gold; bh.font = Font(name="Calibri", size=14, bold=True); bh.alignment = _CTR
    bh = cell(f"K{box2_hdr}", "Annual"); bh.fill = gold; bh.font = Font(name="Calibri", size=14, bold=True); bh.alignment = _CTR
    box_label(f"I{box2_save}", "Saving per month with OCI", gold)
    c = cell(f"J{box2_save}", f"=D{tr}-J{tr}"); c.fill = gold; c.number_format = _ACCT0
    c = cell(f"K{box2_save}", f"=J{box2_save}*12"); c.fill = gold; c.number_format = _ACCT0
    box_label(f"I{box2_pct}", "Total % Savings", gold)
    c = cell(f"J{box2_pct}", f"=IFERROR(J{box2_save}/D{tr},0)"); c.fill = gold; c.number_format = _PCT

    # ---- product-group summary (below the main table) ----
    _cloud_legend(ws, present_groups, first, last if N else first, total_row, summary_start)

    # ---- conditional formatting ----
    if N:
        # "Good" green when savings > 0
        good = DifferentialStyle(font=Font(color="FF" + _C_GOOD_FONT),
                                 fill=PatternFill(bgColor="FF" + _C_GOOD_FILL))
        ws.conditional_formatting.add(
            f"L{first}:L{last}",
            Rule(type="cellIs", operator="greaterThan", formula=["0"], dxf=good))

    # per-group fill on the main-table Product Group cells (the summary tables handle
    # their own group coloring below the table).
    if N:
        for grp, color in _CLOUD_GROUP_COLORS.items():
            dxf = DifferentialStyle(fill=PatternFill(bgColor="FF" + color))
            ws.conditional_formatting.add(
                f"B{first}:B{last}",
                Rule(type="cellIs", operator="equal",
                     formula=[f'"{grp}"'], dxf=dxf))


def _cloud_legend(ws, present_groups, first, last, total_row, start_row):
    """Monthly + annual "Cloud Service Product Group" summary tables, placed BELOW the
    main Product Breakdown table (columns B:H) rather than off to the right."""
    base = Font(name="Calibri", size=14)
    n = len(present_groups)

    # ---- disclaimer banner (B:G) ----
    ws.merge_cells(f"B{start_row}:G{start_row + 1}")
    nb = ws[f"B{start_row}"]
    nb.value = ("Oracle Pricing Proposal. This is not an Official Contract - "
                "For Budgetary Purposes Only")
    nb.fill = _fill(_C_GRAY)
    nb.font = Font(name="Calibri", size=12, bold=True)
    nb.alignment = _CTR_WRAP

    def head(col, r, text, fill, font_color="000000"):
        c = ws[f"{col}{r}"]
        c.value = text
        c.fill = _fill(fill)
        c.font = Font(name="Calibri", size=14, bold=True, color="FF" + font_color)
        c.alignment = _CTR
        return c

    def money(col, r, formula, fmt):
        c = ws[f"{col}{r}"]
        c.value = formula
        c.font = base
        c.number_format = fmt
        return c

    g = _fill(_C_GREEN_HDR)
    sumrng = f"$B${first}:$B${last}"
    fsum = f"$F${first}:$F${last}"
    ksum = f"$K${first}:$K${last}"

    def build_table(hdr_row, label, annual=False, month_lo=None):
        """Columns: B=group, C=% of bill, D=AWS, E=OCI, F=Savings, G=% Savings,
        H=Monthly/Annual label. Returns (lo, tot)."""
        lo = hdr_row + 1
        hi = lo + n - 1
        tot = hi + 1
        head("B", hdr_row, "Cloud Service Product Group", _C_HDR, "FFFFFF")
        head("C", hdr_row, "% of Bill", _C_GREEN_HDR)
        head("D", hdr_row, "AWS Cost", _C_BLUE, "FFFFFF")
        head("E", hdr_row, "OCI Cost", _C_RED2, "FFFFFF")
        head("F", hdr_row, "Savings in $", _C_GREEN_HDR)
        head("G", hdr_row, "% Savings", _C_GREEN_HDR)
        ws.merge_cells(f"H{lo}:H{hi if n else lo}")
        lc = ws[f"H{lo}"]
        lc.value = label
        lc.font = Font(name="Calibri", size=14, bold=True)
        lc.alignment = _CTR
        for i, grp in enumerate(present_groups):
            r = lo + i
            if annual:
                mr = month_lo + i
                c = ws[f"B{r}"]; c.value = f"=B{mr}"
                money("D", r, f"=D{mr}*12", _ACCT0)
                money("E", r, f"=E{mr}*12", _ACCT0)
            else:
                c = ws[f"B{r}"]; c.value = grp
                money("D", r, f"=SUMIF({sumrng},B{r},{fsum})", _ACCT0)
                money("E", r, f"=SUMIF({sumrng},B{r},{ksum})", _ACCT0)
            c.font = Font(name="Calibri", size=14, bold=True)
            money("C", r, f"=IFERROR(D{r}/$D${tot},0)", _PCT)
            money("F", r, f"=D{r}-E{r}", _ACCT0)
            money("G", r, f"=IFERROR(F{r}/D{r},0)", _PCT)
        # total row
        head("B", tot, "Total", _C_GREEN_HDR)
        for col, formula, fmt in (
            ("C", f"=SUM(C{lo}:C{hi})", _PCT),
            ("D", f"=SUM(D{lo}:D{hi})", _MONEY0),
            ("E", f"=SUM(E{lo}:E{hi})", _MONEY0),
            ("F", f"=D{tot}-E{tot}", _MONEY0),
            ("G", f"=IFERROR(F{tot}/D{tot},0)", _PCT)):
            c = money(col, tot, formula, fmt)
            c.fill = g
            c.font = Font(name="Calibri", size=14, bold=True)
        # per-group conditional fill on the group-name cells
        if n:
            for grp, color in _CLOUD_GROUP_COLORS.items():
                dxf = DifferentialStyle(fill=PatternFill(bgColor="FF" + color))
                ws.conditional_formatting.add(
                    f"B{lo}:B{hi}",
                    Rule(type="cellIs", operator="equal",
                         formula=[f'"{grp}"'], dxf=dxf))
        return lo, tot

    m_hdr = start_row + 3
    m_lo, m_tot = build_table(m_hdr, "Monthly", annual=False)
    a_hdr = m_tot + 2
    build_table(a_hdr, "Annual", annual=True, month_lo=m_lo)


_CLOUD_NOTES = [
    ("Carry cost",
     "Costs for services or service descriptions that lack a clear mapping, "
     "have no direct equivalent, use different units of measure compared to "
     "OCI pricing, or that raise questions for needing clarity from Solutions "
     "Engineer are carried over to avoid overestimating potential savings on "
     "the OCI side.\n(*Denoted with asterisk and blue text)"),
    ("Discounting",
     "An OCI discount (set in the app, next to the ramp graph) is applied to the "
     "OCI list pricing: Total - Discounted = Total - List x (1 - discount). At a "
     "0% discount the discounted OCI total equals the figure shown in the app. "
     "Carried-cost and Marketplace items are excluded from the discount."),
    ("Product Grouping",
     "AWS services are grouped into the 11 OCI product groups using Oracle's "
     "AWS->OCI service comparison mapping (e.g. EC2 and Compute Savings Plans "
     "roll up to Compute, S3 and EBS to Storage, RDS to Database)."),
    ("Marketplace + 3rd Party Services",
     "Any AWS Marketplace item or third-party service represents a carried "
     "cost and is assumed to be available in the OCI Marketplace or at the "
     "same price currently charged by the third-party provider.\n"
     "(*Denoted with asterisk and blue text)"),
    ("Oracle Pricing Estimates",
     "Please note that these figures are preliminary estimates only and do "
     "not constitute a final Bill of Materials (BOM). The detailed migration "
     "costs, strategic approach, final discount structure, and precise product "
     "or service mappings to be thoroughly discussed and finalized between the "
     "Oracle account team and the client."),
]


def _cloud_service_mapping_sheet(ws, rows, oci_discount=0.0):
    """Per-line AWS -> OCI mapping detail (one row per priced bill line), so the
    workbook shows exactly how each source service was mapped and priced."""
    ws.sheet_view.showGridLines = True
    widths = {"A": 3.0, "B": 24.0, "C": 34.0, "D": 46.0, "E": 16.0, "F": 16.0,
              "G": 40.0, "H": 16.0, "I": 16.0, "J": 22.0}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    headers = ["Product Group", "AWS Service", "Source SKU / Meter", "Usage",
               "Source Cost", "OCI Product", "OCI Cost", "Savings", "Status"]
    for i, h in enumerate(headers, start=2):  # start col B
        c = ws.cell(row=2, column=i, value=h)
        c.fill = _fill(_C_HDR)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
        c.alignment = _CTR
        c.border = _THIN_BORDER

    # Collapsible outline: the +/- toggle sits on the group HEADER row (above its detail),
    # so each product group can be collapsed away when you don't care about it.
    from openpyxl.worksheet.properties import Outline
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False,
                                            showOutlineSymbols=True)

    def grp(r):
        return _cloud_product_group(r.get("ociServiceCategory"), r.get("sourceService"),
                                    r.get("ociProduct"))

    # Bucket the lines by product group, in the SAME order they appear on the Pricing
    # Overview; within a group, cluster by AWS service (biggest-spend service first) and
    # order each service's lines by descending source cost so it's easy to follow.
    order_index = {g: i for i, g in enumerate(_CLOUD_GROUP_ORDER)}
    buckets = {}
    for row in rows:
        buckets.setdefault(grp(row), []).append(row)
    ordered_groups = sorted(buckets.keys(), key=lambda g: order_index.get(g, 99))

    def _write_detail(r, row):
        fsm = row.get("fullServiceMapping") or {}
        action = (row.get("costAction") or "")
        src = float(row.get("sourceMonthlyCost") or 0)
        oci = float(row.get("monthly") or 0)
        if action == "remove":
            status = "REMOVED"
        elif action == "carry":
            status = "CARRIED OVER"
        elif row.get("mappingFlag"):
            status = str(row.get("mappingFlag"))
        else:
            status = "Mapped"
        oci_prod = row.get("ociProduct") or fsm.get("ociProduct") or "Needs review"
        qty = fsm.get("quantity")
        usage = (f"{qty:,.0f} {fsm.get('unit','')}".strip()
                 if isinstance(qty, (int, float)) and qty else "")
        vals = [grp(row), row.get("sourceService") or "", fsm.get("sourceProduct") or "",
                usage, round(src, 2), oci_prod, round(oci, 2), round(src - oci, 2), status]
        for i, v in enumerate(vals, start=2):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name="Calibri", size=11)
            c.border = _THIN_BORDER
            if i in (6, 8, 9):
                c.number_format = MONEY2
        if oci_prod == "Needs review":
            ws.cell(row=r, column=7).font = Font(name="Calibri", size=11, color="FFC00000", bold=True)
        # Level-1 detail row -> collapses under its group header. Hidden by default so the
        # sheet opens with every group collapsed (expand the ones you care about).
        rd = ws.row_dimensions[r]
        rd.outline_level = 1
        rd.hidden = True
        return src, oci, (action != "remove")

    r = 3
    src_total = oci_total = 0.0
    for g in ordered_groups:
        grows = buckets[g]
        g_src = sum(float(x.get("sourceMonthlyCost") or 0) for x in grows)
        g_oci = sum(float(x.get("monthly") or 0) for x in grows
                    if (x.get("costAction") or "") != "remove")
        # ---- group header row (level 0; carries the collapse button) ----
        color = _CLOUD_GROUP_COLORS.get(g, _C_HDR)
        hdr_fill = _fill(color)
        gh = ws.cell(row=r, column=2, value=f"{g}  ({len(grows)} lines)")
        gh.font = Font(name="Calibri", size=12, bold=True)
        for i in range(2, 11):
            c = ws.cell(row=r, column=i)
            c.fill = hdr_fill
            c.border = _THIN_BORDER
            if not c.font or not c.font.bold:
                c.font = Font(name="Calibri", size=12, bold=True)
        ws.cell(row=r, column=6, value=round(g_src, 2)).number_format = MONEY2
        ws.cell(row=r, column=8, value=round(g_oci, 2)).number_format = MONEY2
        ws.cell(row=r, column=9, value=round(g_src - g_oci, 2)).number_format = MONEY2
        # Header carries the collapse button (summaryBelow=False); mark it collapsed so the
        # sheet opens with the group folded away.
        ws.row_dimensions[r].collapsed = True
        r += 1
        # ---- detail rows, clustered by AWS service (biggest spend first) ----
        svc_total = {}
        for x in grows:
            svc_total[x.get("sourceService") or ""] = svc_total.get(x.get("sourceService") or "", 0.0) + float(x.get("sourceMonthlyCost") or 0)
        grows_sorted = sorted(
            grows,
            key=lambda x: (-svc_total.get(x.get("sourceService") or "", 0.0),
                           x.get("sourceService") or "",
                           -float(x.get("sourceMonthlyCost") or 0)),
        )
        for row in grows_sorted:
            s, o, counted = _write_detail(r, row)
            if counted:
                src_total += s
                oci_total += o
            r += 1

    # Totals row
    tot = ws.cell(row=r, column=2, value="Total (excl. removed)")
    tot.font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
    tot.fill = _fill(_C_TOTAL)
    for i in range(3, 11):
        c = ws.cell(row=r, column=i)
        c.fill = _fill(_C_TOTAL)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
        c.border = _THIN_BORDER
    ws.cell(row=r, column=6, value=round(src_total, 2)).number_format = MONEY2
    ws.cell(row=r, column=8, value=round(oci_total, 2)).number_format = MONEY2
    ws.cell(row=r, column=9, value=round(src_total - oci_total, 2)).number_format = MONEY2
    for i in (6, 8, 9):
        c = ws.cell(row=r, column=i)
        c.fill = _fill(_C_TOTAL)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
    ws.freeze_panes = "B3"


def _cloud_notes_sheet(ws):
    ws.column_dimensions["A"].width = 86.71
    ws.column_dimensions["B"].width = 45.71
    for coord, text in (("A1", "AWS Service"), ("B1", "Notes")):
        c = ws[coord]
        c.value = text
        c.fill = _fill(_C_HDR)
        c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
        c.alignment = _CTR
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for i, (label, note) in enumerate(_CLOUD_NOTES, start=2):
        a = ws[f"A{i}"]
        a.value = label
        a.font = Font(name="Calibri", size=14, bold=True)
        a.alignment = Alignment(vertical="top")
        b = ws[f"B{i}"]
        b.value = note
        b.font = Font(name="Calibri", size=14)
        b.alignment = wrap_top


def _util_by_year(ramp):
    """Translate the app ramp into 5 yearly utilization fractions (0..1)."""
    if not ramp:
        return [1.0] * 5
    if isinstance(ramp, dict) and ramp.get("utilByYear"):
        vals = [float(v) for v in ramp["utilByYear"]]
    else:
        ceiling = float((ramp or {}).get("ceiling") or 0)
        monthly = [float(x) for x in (ramp or {}).get("monthly") or []]
        if ceiling <= 0 or not monthly:
            return [1.0] * 5
        vals = []
        for y in range(5):
            chunk = monthly[y * 12:(y + 1) * 12]
            vals.append((sum(chunk) / len(chunk) / ceiling) if chunk else 1.0)
    vals = [max(0.0, min(1.0, v)) for v in vals]
    while len(vals) < 5:
        vals.append(1.0)
    return vals[:5]


def build_workbook_bytes(servers, ramp=None, existing_infra_cost=0, shape=None, hide_windows=False, hours=HOURS, bom_name="", auto=False, existing_label="Existing Infra Cost (enter):", oci_discount=0.0, extra_services=None):
    shape = _resolve_shape(shape)
    # Sheet name auto-fills with the chosen shape (e.g. "BOM w E6 Ax"); in
    # processor-matching mode each workload gets its own best-fit shape, so say that.
    sheet_name = ("BOM w Matched Shapes" if auto else f"BOM w {shape['shortLabel']}")[:31]
    wb = Workbook()
    bom = wb.active
    bom.title = sheet_name
    build_bom_sheet(bom, servers, shape, hide_windows, hours)
    # App-added OCI services: price them and add both an itemized sheet and their totals.
    # Split into native OCI (eligible for the OCI discount) and 3rd-party licensing (not).
    extra_oci = extra_tp = 0.0
    if extra_services:
        import oci_catalog
        priced, _ = oci_catalog.price_extras(extra_services, hours)
        for s in priced:
            if s.get("thirdParty"):
                extra_tp += float(s["monthly"] or 0)
            else:
                extra_oci += float(s["monthly"] or 0)
        _build_extra_services_sheet(wb, priced, extra_oci + extra_tp)
    overview = wb.create_sheet("Overview")
    build_overview_sheet(overview, _util_by_year(ramp), float(existing_infra_cost or 0), sheet_name, existing_label, oci_discount, round(extra_oci, 2), round(extra_tp, 2))
    overview.sheet_view.tabSelected = True
    wb.active = wb.sheetnames.index("Overview")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def servers_from_pricing(pricing, raw_rows=None):
    """Build the per-server list the exporter needs from a pricing payload."""
    os_by_id = {}
    if raw_rows:
        for row in raw_rows:
            text = " ".join(str(v) for v in row.values()).lower()
            os_by_id[row.get("__id")] = "windows" if "windows" in text else ("linux" if "linux" in text else "")
    servers = []
    for row in pricing.get("rows", []):
        specs = row.get("specs", {})
        size = row.get("sizeCheck") or {}
        shape_used = row.get("shapeUsed") or {}
        # Per-server shape so each row prices at the rates/SKUs of the shape it
        # was actually mapped to (matters in processor-matching / Auto mode).
        srv_shape = None
        if shape_used.get("computeRate") is not None and shape_used.get("memoryRate") is not None:
            srv_shape = {
                "label": shape_used.get("label") or shape_used.get("shortLabel"),
                "shortLabel": shape_used.get("shortLabel") or shape_used.get("label"),
                "computeSku": shape_used.get("computeSku"),
                "memorySku": shape_used.get("memorySku") or shape_used.get("computeSku"),
                "computeRate": shape_used.get("computeRate"),
                "memoryRate": shape_used.get("memoryRate"),
            }
        servers.append({
            "name": row.get("name") or "Server",
            "ocpus": specs.get("ocpus") or 0,
            "memory": specs.get("memoryGb") or 0,
            "disk": specs.get("blockStorageGb") or 0,
            "os": os_by_id.get(row.get("rowId"), ""),
            "environment": row.get("environment") or "",
            "hours": row.get("hoursPerMonth"),
            "sizeStatus": size.get("status", "ok"),
            "sizeMessage": size.get("message", ""),
            "shape": srv_shape,
        })
    return servers
