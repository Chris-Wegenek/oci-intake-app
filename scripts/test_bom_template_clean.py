"""Regression checks for customer-neutral Full BOM exports."""

import io
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import bom_template


FORBIDDEN = re.compile(
    r"polaris|kohler|uswi|koftp|blazer|updated master inventory|"
    r"\b676\b|113,?368|3\.6 pb|107\.11|911 tb|4,?130|10,?324|"
    r"\bamer\b|current dr documents|warm active-passive|1,?000 key|"
    r"separate sap|sap coexistence|f5-dependent|oracle is willing to invest",
    re.IGNORECASE,
)


def build_test_workbook():
    fields = [
        {"key": "server", "sourceHeader": "Server Name"},
        {"key": "application", "sourceHeader": "Application Name"},
        {"key": "tier", "sourceHeader": "Tier"},
        {"key": "environment", "sourceHeader": "Environment"},
    ]
    rows = [
        {
            "__id": "row-1",
            "server": "CURRENT-APP-01",
            "application": "Current Finance",
            "tier": "Tier 1",
            "environment": "Prod",
        },
        {
            "__id": "row-2",
            "server": "CURRENT-APP-02",
            "application": "Current Finance",
            "tier": "Tier 1",
            "environment": "Test",
        },
    ]
    pricing_rows = []
    for index, row in enumerate(rows, 1):
        pricing_rows.append({
            "rowId": row["__id"],
            "name": row["server"],
            "environment": row["environment"],
            "shapeUsed": {"key": "e6-standard", "shortLabel": "VM.Standard.E6.Flex"},
            "hoursPerMonth": 730,
            "specs": {
                "vcpus": 8,
                "ocpus": 4,
                "memoryGb": 32,
                "blockStorageGb": 256,
                "fileStorageGb": 100 if index == 1 else 0,
            },
            "lineItems": [
                {
                    "sku": "TEST-OCPU",
                    "description": "OCPU hours",
                    "unit": "OCPU-hour",
                    "rate": 0.03,
                    "quantity": 4,
                    "hours": 730,
                    "monthly": 87.60,
                },
                {
                    "sku": "TEST-RAM",
                    "description": "Memory GB hours",
                    "unit": "GB-hour",
                    "rate": 0.002,
                    "quantity": 32,
                    "hours": 730,
                    "monthly": 46.72,
                },
            ],
        })
    pricing = {
        "rows": pricing_rows,
        "totals": {"monthly": 268.64},
        "selectedShape": {
            "key": "e6-standard",
            "shortLabel": "VM.Standard.E6.Flex",
        },
    }
    ramp_monthly = [round(268.64 * month / 12, 2) for month in range(1, 13)]
    data = bom_template.build_full_bom_bytes(
        pricing,
        rows=rows,
        fields=fields,
        ramp={"ceiling": 268.64, "monthly": ramp_monthly},
        bom_name="Current Customer",
        shape={
            "key": "e6-standard",
            "shortLabel": "VM.Standard.E6.Flex",
            "computeSku": "TEST-OCPU",
            "memorySku": "TEST-RAM",
            "computeRate": 0.03,
            "memoryRate": 0.002,
        },
        hours=730,
        block_rate=0.0255,
        vpu_rate=0.0017,
        default_vpus=10,
        file_rate=0.0255,
        windows_rate=0.092,
        windows_sku="TEST-WINDOWS",
        include_diagram=False,
    )
    return load_workbook(io.BytesIO(data), data_only=False)


def main():
    wb = build_test_workbook()

    residue = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and FORBIDDEN.search(value):
                    residue.append(f"{ws.title}!{cell.coordinate}: {value}")
    assert not residue, "Reference customer residue found:\n" + "\n".join(residue[:25])

    apps = wb["Applications Migrated to OCI"]
    assert apps["A7"].value == "Current Finance"
    assert apps["C6"].value == "VM.Standard.E6.Flex Monthly Baseline"
    assert apps["I7"].value == "CURRENT-APP-01, CURRENT-APP-02"
    assert all(apps[f"I{row}"].value is None for row in range(8, 208))
    assert all(apps.row_dimensions[row].hidden for row in range(8, 208))

    storage_headers = [
        wb["Storage"].cell(bom_template.STORAGE_HEADER_ROW, col).value
        for col in range(1, 11)
    ]
    assert storage_headers == [
        "Workload / Service", "Tier", "Environment", "Application", "Source Signal",
        "OCI Storage Target", "Capacity (GB)", "Unit Rate", "Monthly", "Annual",
    ]
    assert wb["Storage"]["A11"].value == "CURRENT-APP-01"
    assert wb["Compute"]["S13"].value == "VM.Standard.E6.Flex Total Annual"

    assert wb["Table of Contents"]["B7"].value == "OCI BOM + Architecture Generator"
    assert wb["Table of Contents"]["B12"].value == "Oracle Cloud Infrastructure"
    assert wb["Assumptions"]["C17"].value.startswith("Used for visual styling only")
    assert wb["Consumption Ramp"]["B12"].value > 0
    assert str(wb["Consumption Ramp"]["C12"].value).startswith("=SUM(")

    for name in ("Networking", "Security KMS", "DR",
                 "Annexure Addendum to Storage"):
        assert wb[name].sheet_state == "hidden", f"{name} should be hidden when empty"
    assert "Other Services" not in wb.sheetnames

    print("Full BOM customer-neutral regression checks passed.")


if __name__ == "__main__":
    main()
