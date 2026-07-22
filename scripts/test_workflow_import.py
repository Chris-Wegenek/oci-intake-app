import os, sys, json, tempfile
os.environ["OCI_APP_NO_BOOTSTRAP"] = "1"
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import bom_export
from openpyxl import Workbook, load_workbook

PASS, FAIL = 0, 0

def check(name, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print("  PASS  %s" % name)
    else:
        FAIL += 1
        print("  FAIL  %s  %s" % (name, detail))

def roundtrip_xlsx(wf):
    wb = Workbook()
    wb.active["A1"] = "sheet"
    bom_export.embed_workflow_state(wb, json.dumps(wf))
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    out = bom_export.read_workflow_state(path)
    os.unlink(path)
    return out

def roundtrip_json(wf):
    fd, path = tempfile.mkstemp(suffix=".json")
    os.close(fd)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(wf, fh)
    out = bom_export.read_workflow_state(path)
    os.unlink(path)
    return out

print("=" * 70)
print("EDGE CASE 1: full realistic workflow (xlsx round-trip)")
full = {
    "intakeMode": "cloud_bill", "providerHint": "aws", "fullServiceBeta": True,
    "hideGpuPricing": False, "hideWindowsPricing": True, "rightsize": True,
    "auto": True, "autoTier": "best", "hoursPerMonth": 730, "hoursOverride": False,
    "bomName": "Acme Migration", "ociDiscount": 0.15, "oicMessagePacks": 2,
    "selectedShape": "E6", "existingInfraCost": 12000, "crossCloudTopTier": False,
    "extraServices": [{"sku": "FastConnect", "monthly": 500}],
    "diagramOptions": {"primaryRegion": "us-ashburn-1", "splitADs": True,
                        "adSplitResources": {"vms": False, "dbs": True},
                        "enableDr": True, "drRegion": "us-phoenix-1",
                        "drReplicate": {"vms": True, "dbs": True, "object": False}},
    "ramp": {"months": 36, "ceiling": 100000, "points": [[0, 1.0], [12, 0.0]]},
    "fields": [{"key": "source_service"}],
    "rows": [{"__id": "r1", "monthly": 100.5, "sourceService": "EC2"}],
    "shapeOverrides": {}, "costOverrides": {}, "approvedFlags": {"r1": True},
    "hiddenSources": [], "selectedRows": ["r1"], "columnPrefs": {}, "resultSort": "monthly",
}
out = roundtrip_xlsx(full)
check("xlsx round-trip preserves all keys", out == full, "mismatch")

print("\nEDGE CASE 2: empty / minimal workflow")
mini = {"rows": []}
check("minimal xlsx round-trip", roundtrip_xlsx(mini) == mini)
check("minimal json round-trip", roundtrip_json(mini) == mini)

print("\nEDGE CASE 3: large payload (>32767 chars, forces chunking)")
big = {"rows": [{"__id": "r%d" % i, "note": "x" * 200} for i in range(400)]}
bj = json.dumps(big)
check("payload actually exceeds one cell", len(bj) > 32767, "len=%d" % len(bj))
out = roundtrip_xlsx(big)
check("large xlsx round-trip intact", out == big,
      "len_in=%d len_out=%s" % (len(big["rows"]), len(out.get("rows", [])) if out else "None"))

print("\nEDGE CASE 4: unicode / special chars")
uni = {"bomName": "Café — Über \"quotes\" \\ 日本語 \n newline", "rows": [{"__id": "u1"}]}
check("unicode xlsx round-trip", roundtrip_xlsx(uni) == uni)
check("unicode json round-trip", roundtrip_json(uni) == uni)

print("\nEDGE CASE 5: workflow_json is None -> embed writes no sheet")
wb = Workbook(); wb.active["A1"] = "x"
bom_export.embed_workflow_state(wb, None)
check("no _workflow sheet created for None", bom_export.WORKFLOW_SHEET not in wb.sheetnames)

print("\nEDGE CASE 6: xlsx with NO _workflow sheet -> read returns None")
wb = Workbook(); wb.active["A1"] = "plain"
fd, path = tempfile.mkstemp(suffix=".xlsx"); os.close(fd); wb.save(path)
check("read returns None when sheet absent", bom_export.read_workflow_state(path) is None)
os.unlink(path)

print("\nEDGE CASE 7: malformed .json -> raises (caught by endpoint as 500)")
fd, path = tempfile.mkstemp(suffix=".json"); os.close(fd)
with open(path, "w") as fh: fh.write("{not valid json,,,}")
raised = False
try:
    bom_export.read_workflow_state(path)
except Exception:
    raised = True
os.unlink(path)
check("malformed json raises (surfaced to user)", raised)

print("\nEDGE CASE 8: empty .xlsx workflow sheet text -> None")
wb = Workbook(); wb.active["A1"] = "x"
ws = wb.create_sheet(bom_export.WORKFLOW_SHEET); ws["A1"] = "header only, no data rows"
fd, path = tempfile.mkstemp(suffix=".xlsx"); os.close(fd); wb.save(path)
check("empty workflow text -> None", bom_export.read_workflow_state(path) is None)
os.unlink(path)

print("\n" + "=" * 70)
print("RESULT: %d passed, %d failed" % (PASS, FAIL))
sys.exit(1 if FAIL else 0)
